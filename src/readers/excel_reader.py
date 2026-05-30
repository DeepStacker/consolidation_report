"""
Hardened Excel Boundary Ingestion Reader.
Redesigned to eliminate all P0 silent failure risks:
1. Eliminates spacer row early-halts via full-sheet checks.
2. Restores float representation scientific notations for accounts.
3. Automatically re-calculates null formula cells in Python.
4. Appends index suffixes to duplicate column headers.
"""

import openpyxl
from typing import List, Dict, Any
from src.models.exceptions import IngestionException, MissingSheetException

def compile_merged_cell_cache(ws: openpyxl.worksheet.worksheet.Worksheet) -> Dict[tuple[int, int], Any]:
    """
    Pre-compiles an in-memory coordinate grid cache mapping all coordinates in merged ranges to their top-left value.
    
    Args:
        ws: The openpyxl worksheet instance.
        
    Returns:
        A dictionary mapping (row, col) coordinate tuples to their top-left merged value.
    """
    coordinate_cache = {}
    for merged_range in ws.merged_cells.ranges:
        # Extract the value from the top-left cell of the range
        top_left_val = ws.cell(row=merged_range.min_row, column=merged_range.min_col).value
        
        # Populate all cells in the range with the top-left value
        for r in range(merged_range.min_row, merged_range.max_row + 1):
            for c in range(merged_range.min_col, merged_range.max_col + 1):
                coordinate_cache[(r, c)] = top_left_val
                
    return coordinate_cache

def get_cell_value_cached(ws: openpyxl.worksheet.worksheet.Worksheet, row: int, col: int, 
                           cache: Dict[tuple[int, int], Any]) -> Any:
    """
    Retrieves a cell value using constant O(1) time coordinate lookups, preserving integer casts for floats.
    
    Args:
        ws: The openpyxl worksheet instance.
        row: Target row index.
        col: Target column index.
        cache: Compiled merged-cell coordinate cache.
        
    Returns:
        The cell value (top-left if merged), casting exact floats to integers.
    """
    # Direct cache hit
    if (row, col) in cache:
        val = cache[(row, col)]
    else:
        # Standard cell lookup
        val = ws.cell(row=row, column=col).value

    # P0 Precision Fix: Cast precise float representations back to integers 
    # to prevent float string scientific notation corruption (e.g. 2.848e14)
    if isinstance(val, float) and not val.is_integer() and str(val).endswith(".0"):
        val = int(val)
    elif isinstance(val, float) and val.is_integer():
        val = int(val)
        
    return val

def is_totals_row(row_vals: List[Any], raw_headers: List[str]) -> bool:
    """
    Production-safe totals detector that distinguishes real totals summaries from valid data rows.
    
    A totals row has:
    1. The first cell containing total-terminology (e.g. 'Total', 'Grand Total', 'Final Sum').
    OR
    2. Essential text columns (Assayer Name, Code, BRANCH, or SOL ID) are completely empty,
       but numerical payment or fees columns are populated.
       
    Args:
        row_vals: The extracted list of row values.
        raw_headers: The header names of the sheet.
        
    Returns:
        True if the row is a totals summary block, False otherwise.
    """
    # 1. Terminal word check across all cells in the row
    for val in row_vals:
        if val is not None and isinstance(val, str) and any(kw in val.lower() for kw in ["total", "summary", "final sum", "total visit"]):
            return True
        
    # 2. Critical text fields vs. Numerical counts check (Guruprasad Shet vs Axis MD Row 13 resolver)
    headers_lower = [h.lower() for h in raw_headers]
    
    # Identify indices of critical text fields
    text_fields = ["assayer name", "assayer code", "branch", "branch name", "sol id", "payee code", "name of auditor"]
    critical_indices = [idx for idx, h in enumerate(headers_lower) if any(tf in h for tf in text_fields)]
    
    # Identify indices of numeric total columns
    numeric_fields = ["total pay", "final client fees", "total", "auditor payment", "base audit fee", "total pay (base)"]
    numeric_indices = [idx for idx, h in enumerate(headers_lower) if any(nf in h for nf in numeric_fields)]
    
    if critical_indices and numeric_indices:
        critical_vals = [row_vals[idx] for idx in critical_indices]
        numeric_vals = [row_vals[idx] for idx in numeric_indices]
        
        # If all critical text cells are blank/None, but we have numbers in fee columns, it is a Totals Row!
        all_text_empty = all(v is None or (isinstance(v, str) and v.strip() == "") for v in critical_vals)
        has_numeric_totals = any(isinstance(v, (int, float)) and v > 0 for v in numeric_vals)
        
        if all_text_empty and has_numeric_totals:
            return True
            
    return False

def recalculate_formulas_fallback(row_dict: Dict[str, Any]) -> Dict[str, Any]:
    """
    Safe formula engine.
    If calculated columns return None due to uncalculated/manual formula sheets, 
    re-computes the values programmatically in Python.
    
    Args:
        row_dict: The mapped canonical row dictionary.
        
    Returns:
        The dictionary with validated numeric calculations.
    """
    # 1. Total pay (Base) calculation fallback
    visits = row_dict.get("No. of Visits")
    fee = row_dict.get("Base Audit Fee")
    base_pay = row_dict.get("Total pay (Base)")
    
    # Convert types safely if present
    v_num = int(visits) if visits is not None else 0
    f_num = float(fee) if fee is not None else 0.0
    
    if base_pay is None or (isinstance(base_pay, (int, float)) and base_pay == 0.0 and v_num > 0 and f_num > 0.0):
        row_dict["Total pay (Base)"] = float(v_num) * f_num

    # 2. Total Net pay calculation fallback
    base = row_dict.get("Total pay (Base)", 0.0)
    travel = row_dict.get(" Travel charges(If any)", 0.0)
    cancel = row_dict.get("Branch Cancellation Charges", 0.0)
    an_exp = row_dict.get(" Andaman & Nicobar Branch Expenses", 0.0)
    err = row_dict.get("Error Deduction", 0.0)
    net_pay = row_dict.get("Total pay")
    
    # Safely extract floats
    base_val = float(base) if base is not None else 0.0
    travel_val = float(travel) if travel is not None else 0.0
    cancel_val = float(cancel) if cancel is not None else 0.0
    an_val = float(an_exp) if an_exp is not None else 0.0
    err_val = float(err) if err is not None else 0.0
    
    if net_pay is None or (isinstance(net_pay, (int, float)) and net_pay == 0.0 and base_val > 0.0):
        row_dict["Total pay"] = base_val + travel_val + cancel_val + an_val - err_val
        
    return row_dict

def ingest_raw_rows(filepath: str, sheet_name: str, header_row: int = 1, 
                    data_start_row: int = 2) -> List[Dict[str, Any]]:
    """
    Loads an Excel sheet, pre-compiles cache coordinates, and parses records up to max_row, skipping blank rows.
    
    Args:
        filepath: Absolute filesystem path to the Excel file.
        sheet_name: Target sheet to load.
        header_row: 1-based index where headers are located.
        data_start_row: 1-based index where records begin.
        
    Returns:
        A list of row dictionaries mapping suffix-protected headers to cell values.
        
    Raises:
        IngestionException: On file load or reading failure.
        MissingSheetException: If target sheet is absent.
    """
    try:
        # Load workbook in data_only mode to read evaluated formula results
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        raise IngestionException(f"Failed to open workbook file at {filepath}: {e}")

    if sheet_name not in wb.sheetnames:
        raise MissingSheetException(f"Required sheet '{sheet_name}' is missing in workbook: {filepath}")

    ws = wb[sheet_name]
    max_col = ws.max_column
    max_row = ws.max_row

    # Compile the high-performance merged cell coordinate lookup cache
    cache = compile_merged_cell_cache(ws)

    # P0 Suffix Protection: Renames duplicate column headers dynamically (e.g. Date -> Date, Date_1)
    seen_headers = {}
    raw_headers = []
    for c in range(1, max_col + 1):
        header_val = ws.cell(row=header_row, column=c).value
        h = str(header_val).strip() if header_val is not None else f"Column_{c}"
        if h in seen_headers:
            seen_headers[h] += 1
            h_new = f"{h}_{seen_headers[h]}"
        else:
            seen_headers[h] = 0
            h_new = h
        raw_headers.append(h_new)

    clean_records = []
    
    for r in range(data_start_row, max_row + 1):
        # Extract row cells resolving merged coordinates
        row_vals = [get_cell_value_cached(ws, r, c, cache) for c in range(1, max_col + 1)]
        
        # P0 Empty-Row protection: Filter out completely blank rows
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row_vals):
            continue
            
        # P0 Totals-Detection: Filter out Totals rows cleanly
        if is_totals_row(row_vals, raw_headers):
            continue
            
        # Build raw dictionary map
        row_dict = {}
        for c_idx, h in enumerate(raw_headers):
            row_dict[h] = row_vals[c_idx]
            
        # Apply formula fallback layer (re-calculates uncalculated None formulas)
        row_dict = recalculate_formulas_fallback(row_dict)
            
        clean_records.append(row_dict)

    wb.close()
    return clean_records
