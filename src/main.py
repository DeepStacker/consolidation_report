import os
import re
import sys
import glob
import shutil
import pandas as pd
from typing import Dict, Any, List, Tuple

from src.schema_loader import load_schema_config
from src.readers.excel_reader import ingest_raw_rows
from src.mappers.structure_mapper import map_raw_to_canonical
from src.validators.format_validator import validate_sheet
from src.rules.overrides import global_rules_engine, try_ingest_client_fallback
from src.reconciliation.engine import PipelineReconciler
from src.run_logging.logger import RunAuditLogger
from src.writers.excel_writer import write_consolidated_workbook
from src.models.domain_models import SchemaDefinition
from src.models.exceptions import ConsolidationPlatformException


def discover_clients(config_dir: str) -> List[Tuple[str, SchemaDefinition]]:
    schemas = []
    for f in os.listdir(config_dir):
        if f.endswith(".yaml") or f.endswith(".yml"):
            path = os.path.join(config_dir, f)
            try:
                schema = load_schema_config(path)
                if schema.active:
                    schemas.append((path, schema))
            except Exception as e:
                print(f"  [Warning] Skipping schema {f}: {e}")
    schemas.sort(key=lambda s: s[1].client_id)
    return schemas


def _normalize_string(s: str) -> str:
    """Normalize a string by converting to lowercase, removing file extension,
    replacing non-alphanumeric characters with spaces, and stripping."""
    if not s:
        return ""
    import re
    s_lower = s.lower()
    if s_lower.endswith((".xlsx", ".xls")):
        s = s[:-5] if s_lower.endswith(".xlsx") else s[:-4]
    s = re.sub(r'[^a-z0-9]', ' ', s.lower())
    return " ".join(s.split())


def _compute_filename_similarity(filename: str, schema: SchemaDefinition) -> float:
    """Compute a matching similarity score between 0.0 and 1.0."""
    fn_norm = _normalize_string(filename)
    fn_tokens = set(fn_norm.split())
    if not fn_tokens:
        return 0.0
    
    patterns = []
    if schema.client_id:
        patterns.append(schema.client_id)
    if schema.client_display_name:
        patterns.append(schema.client_display_name)
    if schema.filename_pattern:
        patterns.append(schema.filename_pattern.replace("*", ""))
        
    best_score = 0.0
    for pat in patterns:
        pat_norm = _normalize_string(pat)
        pat_tokens = pat_norm.split()
        if not pat_tokens:
            continue
            
        matches = sum(1 for tok in pat_tokens if tok in fn_tokens)
        overlap_score = matches / len(pat_tokens)
        
        pat_set = set(pat_tokens)
        intersection = fn_tokens.intersection(pat_set)
        union = fn_tokens.union(pat_set)
        jaccard = len(intersection) / len(union) if union else 0.0
        
        score = (overlap_score * 0.75) + (jaccard * 0.25)
        if score > best_score:
            best_score = score
            
    return best_score


def _quick_column_similarity(filepath: str, schema: SchemaDefinition) -> float:
    """Read the first sheet's headers from *filepath* and score against *schema*'s canonicals."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        headers = []
        if wb.sheetnames:
            ws = wb[wb.sheetnames[0]]
            for cell in next(ws.iter_rows(max_row=1, values_only=True), []):
                if cell is not None:
                    h = str(cell).lower().strip()
                    if h:
                        headers.append(h)
        wb.close()
    except Exception:
        return 0.0
    if not headers:
        return 0.0

    schema_canonicals = set()
    for sdef in schema.sheets.values():
        for col in sdef.columns:
            if col.canonical_name:
                schema_canonicals.add(col.canonical_name.lower())

    if not schema_canonicals:
        return 0.0

    fh_matches = 0
    for h in headers:
        if h in schema_canonicals:
            fh_matches += 1
            continue
        norm_h = re.sub(r'[^a-z0-9]', '', h)
        for sc in schema_canonicals:
            if re.sub(r'[^a-z0-9]', '', sc) == norm_h:
                fh_matches += 1
                break

    sc_matches = 0
    for sc in schema_canonicals:
        norm_sc = re.sub(r'[^a-z0-9]', '', sc)
        for h in headers:
            if re.sub(r'[^a-z0-9]', '', h) == norm_sc:
                sc_matches += 1
                break

    file_coverage = fh_matches / max(len(headers), 1)
    schema_coverage = sc_matches / max(len(schema_canonicals), 1)
    return file_coverage * 0.6 + schema_coverage * 0.4


def find_client_file(workspace_path: str, schema: SchemaDefinition, claimed_files: set | None = None) -> str:
    """Find a workbook file matching *schema*.
    
    ``claimed_files`` — set of basenames already assigned to other schemas.
    Files in this set will be skipped to prevent one file being processed by
    multiple schemas.
    """
    if claimed_files is None:
        claimed_files = set()

    def _unclaimed(fp: str) -> bool:
        return os.path.basename(fp) not in claimed_files

    # Tier 1: Try glob pattern first (exact discovery)
    pattern = os.path.join(workspace_path, schema.filename_pattern)
    matches = [fp for fp in glob.glob(pattern) if _unclaimed(fp)]
    if not matches:
        # Tier 2: Case-insensitive substring matching
        all_files = os.listdir(workspace_path)
        pattern_lower = schema.filename_pattern.replace("*", "").lower()
        matches = [
            os.path.join(workspace_path, f) for f in all_files
            if _unclaimed(os.path.join(workspace_path, f))
            and f.endswith((".xlsx", ".xls")) and pattern_lower in f.lower()
        ]
    
    if not matches:
        # Tier 3: Fuzzy token similarity matching
        all_files = os.listdir(workspace_path)
        candidates = []
        for f in all_files:
            fp = os.path.join(workspace_path, f)
            if not _unclaimed(fp) or not f.endswith((".xlsx", ".xls")) or f == "Consolidated_Report.xlsx":
                continue
            
            sim = _compute_filename_similarity(f, schema)
            if sim >= 0.5:
                candidates.append((sim, fp))
        
        candidates.sort(key=lambda x: -x[0])
        matches = [c[1] for c in candidates]
        
    if not matches:
        # Tier 4: Structural sheet-name match fallback
        all_files = os.listdir(workspace_path)
        structural_matches = []
        schema_sheets = set(schema.sheets.keys())
        
        # Load all other active schemas for comparison to prevent "stealing" of files
        all_active_schemas = []
        config_dir = os.path.join(workspace_path, "config", "schemas")
        if os.path.exists(config_dir):
            for f_yaml in os.listdir(config_dir):
                if f_yaml.endswith((".yaml", ".yml")):
                    try:
                        other_schema = load_schema_config(os.path.join(config_dir, f_yaml))
                        if other_schema.active:
                            all_active_schemas.append(other_schema)
                    except Exception:
                        continue
                        
        for f in all_files:
            fp = os.path.join(workspace_path, f)
            if not _unclaimed(fp) or not f.endswith((".xlsx", ".xls")) or f == "Consolidated_Report.xlsx":
                continue
            
            # Compute similarity
            sim = _compute_filename_similarity(f, schema)
            
            # Check if there is another schema that is a significantly better filename match
            better_match_exists = False
            for other in all_active_schemas:
                if other.client_id == schema.client_id:
                    continue
                other_sim = _compute_filename_similarity(f, other)
                if other_sim > sim and other_sim >= 0.3:
                    better_match_exists = True
                    break
            
            if better_match_exists:
                continue
                
            try:
                import openpyxl
                wb = openpyxl.load_workbook(fp, read_only=True)
                sheets = set(wb.sheetnames)
                wb.close()
                overlap = schema_sheets.intersection(sheets)
                if len(overlap) == len(schema_sheets) and len(schema_sheets) > 0:
                    structural_matches.append(fp)
            except Exception:
                continue
        if structural_matches:
            matches = structural_matches

    if not matches:
        # Tier 5: Column-header matching — read file headers from first sheet and
        # compare against schema canonical names
        all_files = [f for f in os.listdir(workspace_path) if f.endswith((".xlsx", ".xls")) and f != "Consolidated_Report.xlsx"]
        schema_canonicals = set()
        for sdef in schema.sheets.values():
            for col in sdef.columns:
                if col.canonical_name:
                    schema_canonicals.add(col.canonical_name.lower())

        if schema_canonicals:
            candidates = []
            for f in all_files:
                fp = os.path.join(workspace_path, f)
                if not _unclaimed(fp):
                    continue
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
                    headers = []
                    if wb.sheetnames:
                        ws = wb[wb.sheetnames[0]]
                        for cell in next(ws.iter_rows(max_row=1, values_only=True), []):
                            if cell is not None:
                                h = str(cell).lower().strip()
                                if h:
                                    headers.append(h)
                    wb.close()
                except Exception:
                    continue
                if not headers:
                    continue

                fh_matches = 0
                for h in headers:
                    if h in schema_canonicals:
                        fh_matches += 1
                        continue
                    norm_h = re.sub(r'[^a-z0-9]', '', h)
                    for sc in schema_canonicals:
                        if re.sub(r'[^a-z0-9]', '', sc) == norm_h:
                            fh_matches += 1
                            break

                sc_matches = 0
                for sc in schema_canonicals:
                    norm_sc = re.sub(r'[^a-z0-9]', '', sc)
                    for h in headers:
                        if re.sub(r'[^a-z0-9]', '', h) == norm_sc:
                            sc_matches += 1
                            break

                file_coverage = fh_matches / max(len(headers), 1)
                schema_coverage = sc_matches / max(len(schema_canonicals), 1)
                score = file_coverage * 0.6 + schema_coverage * 0.4

                if score >= 0.4:
                    candidates.append((score, fp))

            candidates.sort(key=lambda x: -x[0])
            if candidates:
                matches = [c[1] for c in candidates]

    if not matches:
        raise FileNotFoundError(
            f"Could not find the workbook for '{schema.client_display_name}'.\n"
            f"Expected a filename matching: {schema.filename_pattern}\n"
            f"Make sure the file is in the same folder as this program."
        )
    # Prefer files that don't have " 2" suffix
    best = matches[0]
    for m in matches:
        base = os.path.basename(m)
        if " 2.xlsx" not in base:
            best = m
            break
    return best



def execute_e2e_consolidation(workspace_path: str, output_path: str, manual_mappings: dict | None = None):
    """Run the full consolidation pipeline.
    
    ``manual_mappings`` — optional ``{filename: schema_client_id}`` dict that
    overrides auto-detection and forces a specific file → schema binding.
    """
    print("=" * 80)
    print("DYNAMIC CONSOLIDATION PIPELINE")
    print("=" * 80)

    logger = RunAuditLogger()
    logger.log_rule("INIT", "Dynamic consolidation pipeline initialized.")

    config_dir = os.path.join(workspace_path, "config", "schemas")
    if not os.path.exists(config_dir):
        err_msg = f"Schema config directory not found: {config_dir}"
        logger.finalize("FAILED", error=Exception(err_msg))
        logger.write_log(workspace_path)
        raise FileNotFoundError(err_msg)

    try:
        # 1. Discover and load all active schemas
        print("\n[Step 1] Discovering client schemas...")
        discovered = discover_clients(config_dir)
        if not discovered:
            raise FileNotFoundError(f"No active schemas found in {config_dir}")
        print(f"  Found {len(discovered)} active client schema(s).")
        for path, schema in discovered:
            print(f"    - {schema.client_id} ({schema.client_display_name})")

        # Build reverse lookup: filename → schema_client_id for manual mappings
        manual_by_schema: dict[str, str] = {}
        if manual_mappings:
            for fn, cid in manual_mappings.items():
                manual_by_schema[cid] = fn

        # 2. Pre‑match: for each schema, compute its best column‑header similarity
        # against unclaimed files.  This ensures schemas are tried in order of
        # genuine content similarity, not alphabetical order.
        print("\n[Step 2] Pre‑computing file‑schema match scores...")
        schema_match_scores: list[tuple[float, str, SchemaDefinition]] = []
        for schema_path, schema in discovered:
            if schema.client_id in manual_by_schema:
                # Manually assigned schemas always go first (score = 999)
                schema_match_scores.append((999.0, schema.client_id, schema))
                continue
            try:
                # Quick score: try filename tiers 1-3 first (no I/O)
                all_fnames = [f for f in os.listdir(workspace_path) if f.endswith((".xlsx", ".xls")) and f != "Consolidated_Report.xlsx"]
                best = 0.0
                pattern = schema.filename_pattern.replace("*", "").lower() if schema.filename_pattern else ""
                for fname in all_fnames:
                    fn_base = fname.lower()
                    # Tier 1/2: pattern substring
                    if pattern and pattern in fn_base:
                        best = max(best, 1.0)
                    # Tier 3: fuzzy token
                    sim = _compute_filename_similarity(fname, schema)
                    best = max(best, sim)
                    # Tier 5: column‑header (read file only if needed)
                    if best < 0.4:
                        col_score = _quick_column_similarity(os.path.join(workspace_path, fname), schema)
                        best = max(best, col_score)
                schema_match_scores.append((best, schema.client_id, schema))
            except Exception:
                schema_match_scores.append((0.0, schema.client_id, schema))
        # Highest score first → best‑matching schemas claim their files earliest
        schema_match_scores.sort(key=lambda x: (-x[0], x[1]))

        # 3. Process each client: find file, ingest, map, apply rules
        print("\n[Step 2] Processing client workbooks...")
        claimed_files: set[str] = set()  # basenames already assigned to a schema
        client_data: Dict[str, Dict[str, List[Dict[str, Any]]]] = {}
        all_schemas = []
        processed_any = False

        for score, cid, schema in schema_match_scores:
            all_schemas.append(schema)
            print(f"\n  Processing: {schema.client_id}...")

            filepath = None
            if schema.client_id in manual_by_schema:
                preferred = manual_by_schema[schema.client_id]
                fp = os.path.join(workspace_path, preferred)
                if os.path.isfile(fp):
                    filepath = fp
                    print(f"    Using manually assigned file: {preferred}")
                    logger.log_rule("MANUAL", f"User-assigned file {preferred} -> {schema.client_id}")
                else:
                    print(f"    ⚠ Manually assigned file '{preferred}' not found — trying auto-detect")

            if filepath is None:
                try:
                    filepath = find_client_file(workspace_path, schema, claimed_files=claimed_files)
                except FileNotFoundError:
                    print(f"    ⚠ No matching file found — skipping {schema.client_display_name}")
                    logger.log_rule("SKIP", f"No source file for {schema.client_id}, skipped.")
                    continue

            claimed_files.add(os.path.basename(filepath))
            processed_any = True
            logger.log_file(schema.client_id, filepath)
            print(f"    Source: {os.path.basename(filepath)}")

            client_sheets = {}
            for sheet_name, sheet_def in schema.sheets.items():
                print(f"    Reading sheet: {sheet_name}")
                raw = ingest_raw_rows(
                    filepath, sheet_name,
                    sheet_def.header_row, sheet_def.data_start_row
                )
                mapped = map_raw_to_canonical(raw, sheet_name, schema)
                transformed = global_rules_engine.execute_rules_on_records(
                    mapped, schema.client_id, sheet_name
                )
                client_sheets[sheet_name] = transformed
                print(f"      Rows: {len(transformed)}")

            client_data[schema.client_id] = client_sheets

        if not processed_any:
            err_msg = "No source files matched any active schema. Nothing to consolidate."
            logger.log_rule("FAIL", err_msg)
            logger.finalize("FAILED", error=FileNotFoundError(err_msg))
            logger.write_log(workspace_path)
            raise FileNotFoundError(err_msg)

        # 3. Client-specific fallbacks (e.g. Gold Loan)
        try_ingest_client_fallback(workspace_path, client_data, logger)

        # 4. Validate and log counts
        print("\n[Step 3] Validating records...")
        all_validated: Dict[str, List[Dict[str, Any]]] = {}
        all_warnings = []

        for schema in all_schemas:
            for sheet_name, sheet_def in schema.sheets.items():
                records = client_data.get(schema.client_id, {}).get(sheet_name, [])
                validated, warnings = validate_sheet(records, sheet_def)
                all_validated.setdefault(sheet_name, []).extend(validated)
                all_warnings.extend(warnings)
                logger.log_counts(
                    f"{schema.client_id} - {sheet_name}",
                    0, len(validated)
                )

        for w in all_warnings:
            logger.log_warning(w["field"], w["row_idx"], w["message"])

        dup_warnings = [w for w in all_warnings if w["field"] == "DUPLICATE"]
        if dup_warnings:
            print(f"\nValidation Notice: {len(dup_warnings)} duplicate records found.")
            for d in dup_warnings[:5]:
                print(f"  {d['message']}")

        # 5. Build consolidated DataFrames per sheet (once per sheet)
        print("\n[Step 4] Building consolidated DataFrames...")
        cons_dfs: Dict[str, pd.DataFrame] = {}
        sheets_done = set()
        for schema in all_schemas:
            for sheet_name in schema.sheets:
                if sheet_name in sheets_done or sheet_name not in all_validated:
                    continue
                sheets_done.add(sheet_name)
                cons_dfs[sheet_name] = pd.DataFrame(all_validated[sheet_name])

        # 6. Reconciliation
        print("\n[Step 5] Running reconciliation checks...")
        client_dfs: Dict[str, Dict[str, pd.DataFrame]] = {}
        for schema in all_schemas:
            cid = schema.client_id
            client_dfs[cid] = {}
            for sheet_name in schema.sheets:
                records = client_data.get(cid, {}).get(sheet_name, [])
                client_dfs[cid][sheet_name] = pd.DataFrame(records)

        reconciler = PipelineReconciler(client_dfs)
        sheets_done.clear()
        for schema in all_schemas:
            for sheet_name, sheet_def in schema.sheets.items():
                if sheet_name in sheets_done or sheet_name not in cons_dfs:
                    continue
                sheets_done.add(sheet_name)
                reconciler.verify_sheet_reconciliation(
                    sheet_name, cons_dfs[sheet_name],
                    sum_columns=sheet_def.sum_columns
                )

        logger.finalize("SUCCESS")

        # Log sums from all schemas (once per sheet)
        sheets_done.clear()
        for schema in all_schemas:
            for sheet_name, sheet_def in schema.sheets.items():
                if sheet_name in sheets_done or sheet_name not in cons_dfs:
                    continue
                sheets_done.add(sheet_name)
                for sum_col in sheet_def.sum_columns:
                    if sum_col in cons_dfs[sheet_name].columns:
                        input_sum = 0.0
                        for cid in client_dfs:
                            df = client_dfs[cid].get(sheet_name, pd.DataFrame())
                            if sum_col in df.columns:
                                input_sum += df[sum_col].sum()
                        cons_sum = cons_dfs[sheet_name][sum_col].sum()
                        logger.log_sums(
                            f"{sheet_name} - {sum_col}",
                            input_sum, cons_sum
                        )

        # 7. Backup
        if os.path.exists(output_path):
            backup_path = output_path.replace(".xlsx", "_backup.xlsx")
            print(f"\nSafeguard: Backing up existing file to: {os.path.basename(backup_path)}...")
            shutil.copy2(output_path, backup_path)
            logger.log_rule("SAFEGUARD", f"Backed up original file to {os.path.basename(backup_path)}")

        # 8. Write output
        print("\n[Step 6] Writing consolidated Excel workbook...")
        write_consolidated_workbook(cons_dfs, all_schemas, output_path)
        logger.log_rule("WRITE", f"Saved consolidated workbook at: {os.path.basename(output_path)}")

        logger.write_log(workspace_path)
        print("\n" + "=" * 80)
        print("CONSOLIDATION SUCCESSFUL! 100% RECONCILIATION MATCH VERIFIED.")
        print("=" * 80)

    except Exception as e:
        print(f"\nPipeline aborted: {e}", file=sys.stderr)
        logger.finalize("FAILED", error=e)
        logger.write_log(workspace_path)
        raise e


if __name__ == "__main__":
    current_dir = os.path.dirname(os.path.abspath(__file__))
    workspace_dir = os.path.dirname(current_dir)
    target_xlsx = os.path.join(workspace_dir, "Feb'26 consolidated.xlsx")
    execute_e2e_consolidation(workspace_dir, target_xlsx)
