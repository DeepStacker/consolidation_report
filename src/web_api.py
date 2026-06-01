import os
import shutil
import tempfile
import uuid
import sys
import io
import json
import urllib.request
import asyncio
import time
import contextlib
import contextvars
from io import StringIO
from typing import List, Dict, Any
from datetime import datetime, date
from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import FileResponse, Response
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.gzip import GZipMiddleware
from fastapi.staticfiles import StaticFiles
from src.main import execute_e2e_consolidation
from src.models.domain_models import SchemaDefinition

def _normalize_header(h: str) -> str:
    """Replace embedded newlines with space and collapse whitespace."""
    return ' '.join(h.replace('\r', ' ').replace('\n', ' ').split())


app = FastAPI(title="Consolidation Pipeline API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

app.add_middleware(GZipMiddleware, minimum_size=500)

base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
frontend_dist = os.path.join(base_dir, "frontend", "dist")

# Load .env file manually at startup
env_path = os.path.join(base_dir, ".env")
if os.path.exists(env_path):
    with open(env_path, "r", encoding="utf-8") as env_file:
        for line in env_file:
            line = line.strip()
            if line and not line.startswith("#") and "=" in line:
                key, val = line.split("=", 1)
                os.environ[key.strip()] = val.strip()


# In-memory store for output files (keyed by session UUID)
# File bytes are held in RAM only and deleted immediately after download.
# No data is written to persistent disk outside the ephemeral workspace.
FILE_STORE: Dict[str, dict] = {}  # { file_id: {"bytes": bytes, "audit": dict} }
BATCH_STORE: List[dict] = []  # ordered list of batch records
MAX_FILE_STORE = 50  # max files kept in memory
file_store_lock = asyncio.Lock()
batch_store_lock = asyncio.Lock()

async def _trim_file_store():
    """Remove oldest entries when FILE_STORE exceeds MAX_FILE_STORE."""
    while len(FILE_STORE) > MAX_FILE_STORE:
        oldest = next(iter(FILE_STORE))
        del FILE_STORE[oldest]

# Per-request stdout capture via contextvars (safe for concurrent async requests)
_capture_stream = contextvars.ContextVar('_capture_stream', default=None)

class _ContextAwareStdout:
    def write(self, text):
        stream = _capture_stream.get(None)
        if stream is not None:
            stream.write(text)
        else:
            sys.__stdout__.write(text)
    def flush(self):
        stream = _capture_stream.get(None)
        if stream is not None:
            stream.flush()
        else:
            sys.__stdout__.flush()

sys.stdout = _ContextAwareStdout()

class LogCapture:
    def __init__(self):
        self.stream = StringIO()

    def __enter__(self):
        self.token = _capture_stream.set(self.stream)
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        _capture_stream.reset(self.token)

    def get_logs(self) -> str:
        return self.stream.getvalue()

@app.post("/api/consolidate")
async def consolidate(files: List[UploadFile] = File(...)):
    if not files:
        raise HTTPException(status_code=400, detail="No files uploaded")

    # Create a temporary workspace directory
    tmp_dir = tempfile.mkdtemp(prefix="web_consolidation_")
    try:
        # Copy global config schemas to the temp workspace
        base_dir = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        config_src = os.path.join(base_dir, "config", "schemas")
        if os.path.exists(config_src):
            config_dst = os.path.join(tmp_dir, "config", "schemas")
            os.makedirs(config_dst, exist_ok=True)
            for f in os.listdir(config_src):
                shutil.copy2(os.path.join(config_src, f), os.path.join(config_dst, f))

        # Write uploaded files to temp workspace
        uploaded_files_log = []
        for file in files:
            file_path = os.path.join(tmp_dir, file.filename)
            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            uploaded_files_log.append(file.filename)

        output_path = os.path.join(tmp_dir, "Consolidated_Report.xlsx")
        
        # Execute the pipeline while capturing standard output
        with LogCapture() as capturer:
            print("✓ Preparing workspace...")
            for f in uploaded_files_log:
                print(f"  ✓ {f}")
            print("✓ Running consolidation pipeline...\n")
            
            try:
                execute_e2e_consolidation(tmp_dir, output_path)
                success = True
                error_msg = None
            except Exception as e:
                success = False
                error_msg = str(e)
                print(f"\nERROR: {e}")

        logs = capturer.get_logs()

        if not success:
            async with batch_store_lock:
                BATCH_STORE.insert(0, {
                    "id": str(uuid.uuid4()),
                    "timestamp": datetime.now().isoformat(),
                    "filenames": uploaded_files_log,
                    "file_id": None,
                    "health_score": 0,
                    "status": "FAILED",
                    "error": error_msg,
                })
                if len(BATCH_STORE) > 50:
                    BATCH_STORE.pop()
            return {
                "success": False,
                "logs": logs,
                "error": error_msg
            }

        # Read output file into RAM (never touches persistent disk)
        file_id = str(uuid.uuid4())
        with open(output_path, "rb") as f:
            file_bytes = f.read()

        # Capture audit log data from the ephemeral workspace
        audit_data = {}
        for f in os.listdir(tmp_dir):
            if f.startswith("run_audit_log") and f.endswith(".json"):
                import json
                with open(os.path.join(tmp_dir, f), "r") as log_file:
                    audit_data = json.load(log_file)
                break

        # Parse source file data for in-app comparison
        source_data = {}
        for f in os.listdir(tmp_dir):
            if f.endswith((".xlsx", ".xls")) and f != "Consolidated_Report.xlsx":
                try:
                    import openpyxl
                    src_path = os.path.join(tmp_dir, f)
                    src_wb = openpyxl.load_workbook(src_path, data_only=True, read_only=True)
                    src_sheets = {}
                    for sname in src_wb.sheetnames:
                        sws = src_wb[sname]
                        src_headers = []
                        for cell in next(sws.iter_rows(min_row=1, max_row=1, values_only=True), []):
                            h = _normalize_header(str(cell)) if cell is not None else ""
                            if h:
                                src_headers.append(h)
                        src_rows = []
                        for row in sws.iter_rows(min_row=2, values_only=True):
                            row_data = {}
                            has_data = False
                            for ci, val in enumerate(row):
                                if ci < len(src_headers):
                                    if hasattr(val, 'isoformat'):
                                        val = val.isoformat()
                                    elif not isinstance(val, (str, int, float, bool, type(None))):
                                        val = str(val) if val is not None else None
                                    row_data[src_headers[ci]] = val
                                    if val is not None and val != "":
                                        has_data = True
                            if has_data:
                                src_rows.append(row_data)
                        src_sheets[sname] = {"headers": src_headers, "rows": src_rows}
                    src_wb.close()
                    source_data[f] = src_sheets
                except Exception as e:
                    source_data[f] = {"error": str(e)}

        # Store file bytes + audit data together
        async with file_store_lock:
            FILE_STORE[file_id] = {"bytes": file_bytes, "audit": audit_data, "sources": source_data}
            await _trim_file_store()

        # Create batch record for run history
        batch_record = {
            "id": str(uuid.uuid4()),
            "timestamp": datetime.now().isoformat(),
            "filenames": uploaded_files_log,
            "file_id": file_id,
            "health_score": _compute_audit_summary(audit_data).get("health_score", 100),
            "status": "SUCCESS",
        }
        async with batch_store_lock:
            BATCH_STORE.insert(0, batch_record)  # newest first
            if len(BATCH_STORE) > 50:
                BATCH_STORE.pop()

        # Compute enriched audit summary
        audit_summary = _compute_audit_summary(audit_data)

        return {
            "success": True,
            "logs": logs,
            "file_id": file_id,
            "audit_log": audit_data,
            "audit_summary": audit_summary,
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Pipeline error: {str(e)}")
    finally:
        if os.path.exists(tmp_dir):
            shutil.rmtree(tmp_dir, ignore_errors=True)

@app.get("/api/download/{file_id}")
async def download_file(file_id: str):
    entry = FILE_STORE.get(file_id)
    if entry is None:
        raise HTTPException(status_code=404, detail="File not found or already downloaded")
    file_bytes = entry["bytes"] if isinstance(entry, dict) else entry
    
    return Response(
        content=file_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Consolidated_Report.xlsx"}
    )

def _compute_audit_summary(audit: dict) -> dict:
    """Enrich raw audit log with computed stats and a health score."""
    files = audit.get("files_processed", [])
    row_counts = audit.get("row_counts", {})
    fin_sums = audit.get("financial_sums", {})
    warnings = audit.get("validation_warnings", [])
    status = audit.get("reconciliation_status", "UNKNOWN")

    total_rows = sum(v.get("output", 0) for v in row_counts.values())
    total_sums = len(fin_sums)
    matched_sums = sum(
        1 for v in fin_sums.values()
        if abs(float(v.get("standalone", 0)) - float(v.get("consolidated", 0))) < 0.01
    )

    # Build per-client breakdown
    clients = {}
    for f in files:
        cid = f.get("client_id", "unknown")
        if cid not in clients:
            clients[cid] = {"files": [], "rows": 0, "sheets": set()}
        clients[cid]["files"].append(f.get("filename", ""))
    for key, val in row_counts.items():
        parts = key.split(" - ", 1)
        cid = parts[0].lower() if len(parts) > 1 else "unknown"
        sheet = parts[1] if len(parts) > 1 else key
        if cid in clients:
            clients[cid]["sheets"].add(sheet)
            clients[cid]["rows"] += val.get("output", 0)

    client_list = []
    for cid, data in clients.items():
        client_list.append({
            "client_id": cid,
            "files": data["files"],
            "sheets": sorted(data["sheets"]),
            "total_rows": data["rows"],
        })
    client_list.sort(key=lambda x: x["client_id"])

    # Health score (0-100)
    score = 100
    if status != "SUCCESS":
        score -= 40
    if warnings:
        score -= min(len(warnings) * 3, 30)
    if total_sums > 0 and matched_sums < total_sums:
        score -= 20
    if total_rows == 0 and status == "SUCCESS":
        score -= 20
    score = max(0, score)

    return {
        "status": status,
        "health_score": score,
        "total_files": len(files),
        "total_clients": len(clients),
        "total_rows": total_rows,
        "total_sums": total_sums,
        "matched_sums": matched_sums,
        "total_warnings": len(warnings),
        "clients": client_list,
    }


# ──────────────────────────────────────────────
# SCHEMA CRUD API
# ──────────────────────────────────────────────

import yaml

SCHEMAS_DIR = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "config", "schemas")


def _find_schema_file(client_id: str) -> str | None:
    """Find schema YAML file by client_id."""
    if not os.path.isdir(SCHEMAS_DIR):
        return None
    for f in os.listdir(SCHEMAS_DIR):
        if f.endswith((".yaml", ".yml")):
            fp = os.path.join(SCHEMAS_DIR, f)
            with open(fp) as fh:
                try:
                    data = yaml.safe_load(fh)
                    if isinstance(data, dict) and data.get("client_id") == client_id:
                        return fp
                except Exception:
                    continue
    return None


def _read_schema_yaml(filepath: str) -> dict:
    with open(filepath) as fh:
        return yaml.safe_load(fh)


def _write_schema_yaml(filepath: str, data: dict):
    with open(filepath, "w") as fh:
        yaml.dump(data, fh, default_flow_style=False, sort_keys=False, allow_unicode=True)


@app.get("/api/schemas")
async def list_schemas():
    """List all schema YAML files with summary info."""
    results = []
    if not os.path.isdir(SCHEMAS_DIR):
        return {"schemas": results}
    for f in sorted(os.listdir(SCHEMAS_DIR)):
        if f.endswith((".yaml", ".yml")):
            fp = os.path.join(SCHEMAS_DIR, f)
            try:
                data = _read_schema_yaml(fp)
                sheets = data.get("sheets", {})
                results.append({
                    "filename": f,
                    "client_id": data.get("client_id", ""),
                    "client_display_name": data.get("client_display_name", ""),
                    "filename_pattern": data.get("filename_pattern", ""),
                    "active": data.get("active", True),
                    "sheet_names": list(sheets.keys()),
                    "column_count": sum(len(s.get("columns", [])) for s in sheets.values()),
                })
            except Exception as e:
                results.append({"filename": f, "error": str(e)})
    return {"schemas": results}


@app.get("/api/schemas/all-details")
async def get_all_schemas_details():
    """Get full schema details for all templates in a single batch request."""
    results = []
    if not os.path.isdir(SCHEMAS_DIR):
        return {"schemas": results}
    for f in sorted(os.listdir(SCHEMAS_DIR)):
        if f.endswith((".yaml", ".yml")):
            fp = os.path.join(SCHEMAS_DIR, f)
            try:
                data = _read_schema_yaml(fp)
                results.append(data)
            except Exception as e:
                pass
    return {"schemas": results}


@app.get("/api/schemas/{client_id}")
async def get_schema(client_id: str):
    """Get full schema definition by client_id."""
    fp = _find_schema_file(client_id)
    if not fp:
        raise HTTPException(status_code=404, detail=f"Schema '{client_id}' not found")
    return _read_schema_yaml(fp)


@app.post("/api/schemas")
async def create_schema(data: dict):
    """Create a new schema YAML file."""
    client_id = data.get("client_id", "").strip()
    if not client_id:
        raise HTTPException(status_code=400, detail="client_id is required")
    if _find_schema_file(client_id):
        raise HTTPException(status_code=409, detail=f"Schema '{client_id}' already exists")
    filename = f"{client_id}.yaml"
    filepath = os.path.join(SCHEMAS_DIR, filename)
    os.makedirs(SCHEMAS_DIR, exist_ok=True)
    _write_schema_yaml(filepath, data)
    return {"success": True, "client_id": client_id, "filename": filename}


@app.put("/api/schemas/{client_id}")
async def update_schema(client_id: str, data: dict):
    """Update an existing schema YAML file."""
    fp = _find_schema_file(client_id)
    if not fp:
        raise HTTPException(status_code=404, detail=f"Schema '{client_id}' not found")
    new_id = data.get("client_id", "").strip()
    if new_id and new_id != client_id:
        if _find_schema_file(new_id):
            raise HTTPException(status_code=409, detail=f"Schema '{new_id}' already exists")
        os.remove(fp)
        fp = os.path.join(SCHEMAS_DIR, f"{new_id}.yaml")
    # Validate schema before writing
    try:
        SchemaDefinition.model_validate(data)
    except Exception as e:
        raise HTTPException(422, f"Schema validation failed: {str(e)}")
    _write_schema_yaml(fp, data)
    return {"success": True, "client_id": new_id or client_id}


@app.delete("/api/schemas/{client_id}")
async def delete_schema(client_id: str):
    fp = _find_schema_file(client_id)
    if not fp:
        raise HTTPException(status_code=404, detail=f"Schema '{client_id}' not found")
    os.remove(fp)
    return {"success": True, "client_id": client_id}


@app.put("/api/schemas/{client_id}/toggle")
async def toggle_schema(client_id: str):
    """Toggle active status of a schema."""
    fp = _find_schema_file(client_id)
    if not fp:
        raise HTTPException(status_code=404, detail=f"Schema '{client_id}' not found")
    data = _read_schema_yaml(fp)
    data["active"] = not data.get("active", True)
    _write_schema_yaml(fp, data)
    return {"success": True, "client_id": client_id, "active": data["active"]}


# ──────────────────────────────────────────────
# BATCH HISTORY
# ──────────────────────────────────────────────

@app.get("/api/batches")
async def list_batches():
    """Return all consolidation run batches (newest first)."""
    return {"batches": BATCH_STORE}


@app.delete("/api/batches")
async def clear_batches():
    """Clear all batch history."""
    async with batch_store_lock:
        BATCH_STORE.clear()
    return {"success": True}


# ──────────────────────────────────────────────
# CELL QUALITY ANALYSIS
# ──────────────────────────────────────────────

import re
from collections import Counter


def _infer_column_type(values: list) -> str:
    """Infer the dominant type of a column from sampled values."""
    scores = {"numeric": 0, "date": 0, "code": 0, "text": 0}
    for v in values:
        if v is None or str(v).strip() == "":
            continue
        s = str(v).strip()
        # Phone numbers: treat as text, never as code or numeric
        if re.match(r'^\+?[\d\s\-\(\)]{7,15}$', s) and re.search(r'\d{7,}', s.replace(" ", "").replace("-", "").replace("(", "").replace(")", "")):
            scores["text"] += 1
        elif re.match(r'^\d{1,3}(,\d{3})*(\.\d+)?$', s) or re.match(r'^-?\d+(\.\d+)?$', s):
            scores["numeric"] += 1
        elif re.match(r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}', s) or re.match(r'^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}', s):
            scores["date"] += 1
        elif re.match(r'^[A-Za-z]{2,6}\d{2,10}$', s) or re.match(r'^(?=.*[A-Za-z])(?=.*\d)[A-Za-z0-9]{4,12}$', s):
            scores["code"] += 1
        else:
            scores["text"] += 1
    total = sum(scores.values())
    if total == 0:
        return "unknown"
    best = max(scores, key=scores.get)
    return best if scores[best] / total > 0.5 else "mixed"


def _is_missing(val) -> bool:
    if val is None:
        return True
    s = str(val).strip()
    if s == "" or s == "-" or s == "—" or s == "N/A" or s == "n/a" or s == "NA" or s == "null" or s == "None":
        return True
    if re.match(r'^[Nn][.\s]?[Aa]$', s) or s.lower() in ("nan", "none", "nil", "na"):
        return True
    return False


def _is_date_string(s: str) -> bool:
    return bool(re.match(r'^\d{4}[-/]\d{1,2}[-/]\d{1,2}', s) or re.match(r'^\d{1,2}[-/]\d{1,2}[-/]\d{2,4}', s))


def _is_numeric_string(s: str) -> bool:
    s = s.replace(",", "").replace("%", "").strip()
    try:
        float(s)
        return True
    except ValueError:
        return False


def _get_numeric_value(s: str) -> float | None:
    try:
        return float(s.replace(",", "").replace("%", "").strip())
    except (ValueError, AttributeError):
        return None


def analyze_cell_quality(headers: list, rows: list) -> dict:
    """Return a dict ri -> col -> {"type": str, "message": str} for every problematic cell."""
    issue_map: dict = {}
    if not headers or not rows:
        return issue_map

    # Pre-scan: infer column types and collect numeric values for outlier detection
    col_types = {}
    col_numeric_values: dict[str, list] = {}
    col_format_patterns: dict[str, Counter] = {}
    col_zero_counts: dict[str, int] = {}
    col_non_missing_counts: dict[str, int] = {}

    for ci, h in enumerate(headers):
        values = [row.get(h) for row in rows]
        col_types[h] = _infer_column_type(values)
        # Collect zero and non-missing counts for suspicious-zero detection
        col_zero_counts[h] = sum(1 for v in values if v is not None and str(v).strip() in ("0", "0.0", "0.00"))
        col_non_missing_counts[h] = sum(1 for v in values if v is not None and str(v).strip() != "")
        # Collect numeric values for outlier detection
        if col_types[h] in ("numeric", "mixed"):
            nums = []
            for v in values:
                n = _get_numeric_value(str(v)) if v is not None else None
                if n is not None:
                    nums.append(n)
            if len(nums) >= 4:
                # Check if this is likely a phone/ID column (all values are 7+ digit integers with no decimal)
                # Phone numbers stored as raw integers should NOT get outlier detection
                phone_like_count = sum(1 for n in nums if n >= 10_000_000 and n == int(n) and n == round(n, 0))
                if phone_like_count >= len(nums) * 0.8:
                    continue  # skip phone/ID columns
                # Check if this is a sequential index column (e.g., S.no, serial numbers)
                # Signature: all integers, nearly all unique, range ≈ count-1
                int_count = sum(1 for n in nums if n == int(n) and n == round(n, 0))
                if int_count == len(nums):
                    unique_vals = set(int(n) for n in nums)
                    if len(unique_vals) >= len(nums) * 0.9:
                        mn, mx = min(unique_vals), max(unique_vals)
                        if len(unique_vals) > 1 and (mx - mn) / (len(unique_vals) - 1) >= 0.85:
                            continue  # skip sequential index columns
                col_numeric_values[h] = nums
        # Collect format patterns for text/code columns
        if col_types[h] in ("code", "text"):
            pattern_counts = Counter()
            for v in values:
                if v is not None and str(v).strip():
                    s = str(v).strip()
                    # Create a pattern signature: all letters → "X" (case-insensitive)
                    sig = ""
                    for ch in s:
                        if ch.isalpha():
                            sig += "X"
                        elif ch.isdigit():
                            sig += "0"
                        else:
                            sig += ch
                    pattern_counts[sig] += 1
            if pattern_counts:
                col_format_patterns[h] = pattern_counts

    # Per-row analysis
    for ri, row in enumerate(rows):
        for ci, h in enumerate(headers):
            # Skip columns that come directly from financial institutions (always correct)
            if re.search(r'ifsc', h, re.IGNORECASE):
                continue
            val = row.get(h)
            cell_issues = []

            # 1. Missing check
            if _is_missing(val):
                cell_issues.append(("missing", "Cell is empty"))

            # 2. Pattern check (skip phone-like values to avoid false positives)
            if not _is_missing(val) and col_types.get(h) in ("date", "numeric", "code", "text"):
                s = str(val).strip()
                is_phone_like = re.match(r'^\+?[\d\s\-\(\)]{6,}$', s) and len(re.findall(r'\d', s)) >= 7
                if col_types[h] == "date" and not _is_date_string(s) and not isinstance(val, (date, datetime)):
                    cell_issues.append(("pattern", f"Expected date format, got '{s[:30]}'"))
                elif col_types[h] == "numeric" and not _is_numeric_string(s) and not is_phone_like:
                    cell_issues.append(("pattern", f"Expected numeric value, got '{s[:30]}'"))
                elif col_types[h] in ("code", "text") and not is_phone_like:
                    # Flag suspicious zero values (0 used as placeholder for missing data)
                    if s in ("0", "0.0", "0.00") and col_zero_counts.get(h, 0) < col_non_missing_counts.get(h, 0) * 0.3:
                        cell_issues.append(("pattern", f"Suspicious zero value in {h}, expected text or code"))
                    elif col_types[h] == "text" and re.match(r'^\d+(\.\d+)?$', s) and s not in ("0", "0.0", "0.00"):
                        cell_issues.append(("pattern", f"Numeric value '{s[:20]}' in text column"))
                    elif col_types[h] == "code" and col_format_patterns.get(h):
                        # Skip sentinel values like N.A, N/A, NA, -, etc.
                        if re.match(r'^[Nn][./]?[Aa]$|^[-—]$|^null$|^none$|^nil$|^\.$|^[Nn]a[Nn]$', s):
                            pass
                        else:
                            # Generate normalized pattern sig (all letters → X)
                            sig = "".join("X" if ch.isalpha() else "0" if ch.isdigit() else ch for ch in s)
                            dominant_pattern = col_format_patterns[h].most_common(1)[0][0]
                            dominant_count = col_format_patterns[h][dominant_pattern]
                            total_patterns = sum(col_format_patterns[h].values())
                            if sig != dominant_pattern and dominant_count / total_patterns > 0.6 and total_patterns >= 5:
                                cell_issues.append(("pattern", f"Format '{s[:20]}' differs from column pattern '{dominant_pattern}'"))

            # 3. Outlier check — disabled: too many false positives for financial data
            # (natural variance in payment amounts and counts is legitimate, not anomalous)
            pass

            # 4. Inconsistency check (mixed-format columns)
            if not _is_missing(val) and col_types.get(h) == "mixed":
                s = str(val).strip()
                if _is_numeric_string(s) and not _is_date_string(s):
                    pass  # numeric in mixed column is fine
                elif _is_date_string(s):
                    pass  # date in mixed column is fine
                else:
                    cell_issues.append(("inconsistency", f"Value '{s[:30]}' doesn't match column's dominant types"))

            if cell_issues:
                if ri not in issue_map:
                    issue_map[ri] = {}
                # Keep the most severe issue per cell (missing > outlier > pattern > inconsistency)
                severity = {"missing": 0, "outlier": 1, "pattern": 2, "inconsistency": 3}
                best = min(cell_issues, key=lambda x: severity.get(x[0], 9))
                issue_map[ri][h] = {"type": best[0], "message": best[1]}

    return issue_map


@app.get("/api/preview/{file_id}")
async def preview_file(file_id: str):
    """Return the consolidated Excel data as JSON with cell-level issue highlights."""
    try:
        entry = FILE_STORE.get(file_id)
        if entry is None:
            raise HTTPException(status_code=404, detail="File not found")
        file_bytes = entry["bytes"] if isinstance(entry, dict) else entry
        audit = entry.get("audit", {}) if isinstance(entry, dict) else {}
        sources = entry.get("sources", {}) if isinstance(entry, dict) else {}
        warnings = audit.get("validation_warnings", [])

        import openpyxl
        from io import BytesIO
        wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True, read_only=False)
        sheets = {}
        for name in wb.sheetnames:
            ws = wb[name]
            headers = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []):
                headers.append(_normalize_header(str(cell)) if cell is not None else "")
            rows = []
            issue_map = {}
            for ri, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                row_data = {}
                has_data = False
                for ci, val in enumerate(row):
                    if ci < len(headers):
                        k = headers[ci]
                        # Convert non-serializable types (datetime, date, etc.)
                        if hasattr(val, 'isoformat'):
                            val = val.isoformat()
                        elif not isinstance(val, (str, int, float, bool, type(None))):
                            val = str(val) if val is not None else None
                        row_data[k] = val
                        if val is not None and val != "":
                            has_data = True
                if has_data:
                    row_idx = len(rows)
                    rows.append(row_data)
                    for w in warnings:
                        if w.get("row_idx") == ri:
                            field = w.get("field", "")
                            issue_map.setdefault(row_idx, {})[field] = w.get("message", "")
            has_issues = bool(issue_map)
            sheets[name] = {"headers": headers, "rows": rows, "issues": issue_map, "has_issues": has_issues}

        # Run comprehensive quality analysis on every sheet (merges with existing warnings)
        for name, sheet_data in sheets.items():
            quality_issues = analyze_cell_quality(sheet_data["headers"], sheet_data["rows"])
            existing = sheet_data.get("issues", {})
            for ri, cols in quality_issues.items():
                for col, info in cols.items():
                    if ri not in existing:
                        existing[ri] = {}
                    # Don't overwrite more severe existing issue
                    if col not in existing[ri]:
                        existing[ri][col] = info
                    else:
                        # Merge: if existing is plain string, convert to dict preserving type
                        old = existing[ri][col]
                        if isinstance(old, str):
                            existing[ri][col] = {"type": "warning", "message": old}
            sheet_data["issues"] = existing
            sheet_data["has_issues"] = any(bool(v) for v in existing.values())

        wb.close()

        # Run quality analysis on source file sheets too (backend-driven)
        for fname, fdata in sources.items():
            if isinstance(fdata, dict) and "error" not in fdata:
                for sname, sdata in fdata.items():
                    if isinstance(sdata, dict) and "headers" in sdata and "rows" in sdata:
                        src_quality = analyze_cell_quality(sdata["headers"], sdata["rows"])
                        # Convert int keys to string for JSON serialisation
                        sdata["issues"] = {str(k): v for k, v in src_quality.items()}
                        sdata["has_issues"] = bool(src_quality)

        return {"sheets": sheets, "file_id": file_id, "sources": sources}
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Preview error: {str(e)}")


# ──────────────────────────────────────────────
# PIPELINE PREVIEW & MATCHING
# ──────────────────────────────────────────────

def _normalize_string(s: str) -> str:
    """Normalize a string by converting to lowercase, removing file extension,
    replacing non-alphanumeric characters with spaces, and stripping."""
    if not s:
        return ""
    import re
    s_lower = s.lower()
    if s_lower.endswith((".xlsx", ".xls")):
        s = s[:-5] if s_lower.endswith(".xlsx") else s[:-4]
    s = re.sub(r'[^a-z0-9]', ' ', s.lower())
    return " ".join(s.split())


def _compute_filename_similarity(filename: str, schema: dict) -> float:
    """Compute a matching similarity score between 0.0 and 1.0."""
    fn_norm = _normalize_string(filename)
    fn_tokens = set(fn_norm.split())
    if not fn_tokens:
        return 0.0
    
    patterns = []
    if schema.get("client_id"):
        patterns.append(schema["client_id"])
    if schema.get("client_display_name"):
        patterns.append(schema["client_display_name"])
    if schema.get("filename_pattern"):
        patterns.append(schema["filename_pattern"].replace("*", ""))
        
    best_score = 0.0
    for pat in patterns:
        pat_norm = _normalize_string(pat)
        pat_tokens = pat_norm.split()
        if not pat_tokens:
            continue
            
        matches = sum(1 for tok in pat_tokens if tok in fn_tokens)
        overlap_score = matches / len(pat_tokens)
        
        pat_set = set(pat_tokens)
        intersection = fn_tokens.intersection(pat_set)
        union = fn_tokens.union(pat_set)
        jaccard = len(intersection) / len(union) if union else 0.0
        
        score = (overlap_score * 0.75) + (jaccard * 0.25)
        if score > best_score:
            best_score = score
            
    return best_score


@app.post("/api/preview-matching")
async def preview_matching(data: dict):
    """Check which active schemas match the given filenames (dry-run matching)."""
    filenames = data.get("filenames", [])
    results = {}
    if os.path.isdir(SCHEMAS_DIR):
        for f in os.listdir(SCHEMAS_DIR):
            if not f.endswith((".yaml", ".yml")):
                continue
            try:
                schema = _read_schema_yaml(os.path.join(SCHEMAS_DIR, f))
                if not schema.get("active", True):
                    continue
                pattern = schema.get("filename_pattern", "")
                cid = schema.get("client_id", "")
                display = schema.get("client_display_name", "") or cid
                sheet_names = list(schema.get("sheets", {}).keys())
                sheet_count = len(sheet_names)
            except Exception:
                continue
            for fn in filenames:
                name_lower = fn.lower()
                # Same logic as pipeline's find_client_file
                if not name_lower.endswith((".xlsx", ".xls")):
                    continue
                match = False
                match_score = 0.0
                
                # Tier 1: Exact or pattern substring match
                if pattern:
                    pat_lower = pattern.replace("*", "").lower()
                    if pat_lower in name_lower:
                        match = True
                        match_score = 1.0
                
                # Tier 2: Fuzzy match
                if not match:
                    sim = _compute_filename_similarity(fn, schema)
                    if sim >= 0.5:  # threshold for fuzzy matching
                        match = True
                        match_score = sim
                
                if match:
                    results.setdefault(fn, []).append({
                        "client_id": cid,
                        "client_display_name": display,
                        "filename_pattern": pattern,
                        "sheet_names": sheet_names,
                        "sheet_count": sheet_count,
                        "match_score": match_score,
                    })
    # Sort matches: best match first (highest match score, then most sheets)
    for fn in results:
        results[fn].sort(key=lambda s: (-s["match_score"], -s["sheet_count"]))
    return {"matches": results}



# ──────────────────────────────────────────────
# EXCEL ANALYSIS & CANONICAL FIELDS
# ──────────────────────────────────────────────

import openpyxl


@app.post("/api/preview/{file_id}/save")
async def save_preview(file_id: str, data: dict):
    """Accept edited rows, generate a new Excel file, return a new download file_id."""
    entry = FILE_STORE.get(file_id)
    if entry is None:
        raise HTTPException(status_code=404, detail="File not found")
    orig_bytes = entry["bytes"] if isinstance(entry, dict) else entry

    import openpyxl
    from io import BytesIO
    wb = openpyxl.load_workbook(BytesIO(orig_bytes))

    edited_sheets = data.get("sheets", {})
    for name, sheet_data in edited_sheets.items():
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []):
            headers.append(_normalize_header(str(cell)) if cell is not None else "")

        edited_rows = sheet_data.get("rows", [])
        # Clear existing data rows (keep header)
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row or 2):
            for cell in row:
                cell.value = None

        # Write edited data starting at row 2
        for ri, row_data in enumerate(edited_rows, start=2):
            for ci, h in enumerate(headers):
                if h in row_data:
                    ws.cell(row=ri, column=ci + 1, value=row_data[h])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    new_id = str(uuid.uuid4())
    async with file_store_lock:
        FILE_STORE[new_id] = {"bytes": buf.getvalue(), "audit": entry.get("audit", {}) if isinstance(entry, dict) else {}, "sources": entry.get("sources", {}) if isinstance(entry, dict) else {}}
        await _trim_file_store()
    wb.close()
    return {"success": True, "file_id": new_id}


@app.post("/api/analyze-excel")
async def analyze_excel(file: UploadFile = File(...)):
    """Upload an Excel file and return sheet names, column headers, and preview data rows."""
    if not (file.filename or "").endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Only .xlsx/.xls files are supported")
    tmp = tempfile.mkdtemp(prefix="excel_analyze_")
    try:
        path = os.path.join(tmp, file.filename)
        with open(path, "wb") as f:
            shutil.copyfileobj(file.file, f)
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        sheets = []
        for name in wb.sheetnames:
            ws = wb[name]
            headers = []
            for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True), []):
                if cell is not None:
                    h = _normalize_header(str(cell))
                    if h:
                        headers.append(h)
            # Read up to 5 data rows for preview
            preview = []
            for row in ws.iter_rows(min_row=2, max_row=min(6, ws.max_row or 2), values_only=True):
                row_data = {}
                for i, val in enumerate(row):
                    if i < len(headers):
                        row_data[headers[i]] = val if val is not None else ""
                if any(v != "" for v in row_data.values()):
                    preview.append(row_data)
            sheets.append({"name": name, "columns": headers, "preview": preview})
        wb.close()
        return {"sheets": sheets, "filename": file.filename}
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


@app.get("/api/canonical-fields")
async def canonical_fields():
    """Return union of all canonical field names across all schemas plus defaults."""
    fields = set()
    if os.path.isdir(SCHEMAS_DIR):
        for f in os.listdir(SCHEMAS_DIR):
            if f.endswith((".yaml", ".yml")):
                try:
                    data = _read_schema_yaml(os.path.join(SCHEMAS_DIR, f))
                    for sd in data.get("sheets", {}).values():
                        for col in sd.get("columns", []):
                            if col.get("canonical_name"):
                                fields.add(col["canonical_name"])
                except Exception:
                    continue
    defaults = [
        "S.no", "Sr No", "Client", "Assayer Name", "Assayer Code", "Assayer Phone",
        "Assayer PAN", "Location", "State", "Zone", "Branch", "Branch Code",
        "Month", "Audit Month & Year", "Type of Audit", "No. of Visits",
        "Base Audit Fee", "Total pay (Base)", "Travel charges",
        "Cancelled visits", "Branch Cancellation Charges",
        "Andaman & Nicobar Branch Expenses", "Error Deduction",
        "Total pay", "Remarks", "PAN Number", "Bank Name",
        "A/c Number", "IFSC Code", "Schedule date", "Audit Status",
        "Audit completion date", "No of days audited", "No of Packets audited",
        "Client fee", "Additional", "Final Client Fees", "Assayer fee",
        "Additional fee", "Distance", "Base Location", "Cancelled",
        "Total", "Audit Remarks", "Contact Person", "SOL ID",
    ]
    for d in defaults:
        fields.add(d)
    return {"fields": sorted(fields)}


def get_consolidated_headers(sheet_name: str) -> List[str]:
    filepath = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Feb'26 consolidated.xlsx")
    if not os.path.exists(filepath):
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True)
        target = sheet_name.lower().strip()
        matched = next((name for name in wb.sheetnames if name.lower().strip() == target), None)
        if not matched:
            wb.close()
            return []
        ws = wb[matched]
        row = next(ws.iter_rows(max_row=1, values_only=True), [])
        wb.close()
        return [str(c).strip() for c in row if c is not None and str(c).strip()]
    except Exception:
        return []


def get_consolidated_preview(sheet_name: str, max_rows: int = 20) -> dict:
    """Return headers + up to max_rows of sample data from the consolidated report."""
    filepath = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Feb'26 consolidated.xlsx")
    if not os.path.exists(filepath):
        return {"headers": [], "rows": [], "matched_sheet": ""}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        target = sheet_name.lower().strip()
        matched = next((name for name in wb.sheetnames if name.lower().strip() == target), None)
        if not matched:
            wb.close()
            return {"headers": [], "rows": [], "matched_sheet": ""}
        ws = wb[matched]
        rows_iter = ws.iter_rows(values_only=True)
        header_row = next(rows_iter, [])
        headers = [_normalize_header(str(c)) if c is not None else "" for c in header_row]
        sample_rows = []
        for _ in range(max_rows):
            try:
                row = next(rows_iter)
                row_dict = {}
                for i, val in enumerate(row):
                    if i < len(headers) and headers[i]:
                        if val is not None and str(val).strip():
                            row_dict[headers[i]] = str(val).strip()
                if row_dict:
                    sample_rows.append(row_dict)
            except StopIteration:
                break
        wb.close()
        return {"headers": headers, "rows": sample_rows, "matched_sheet": matched}
    except Exception:
        return {"headers": [], "rows": [], "matched_sheet": ""}


def list_consolidated_sheets() -> List[str]:
    """Return all sheet names in the consolidated report."""
    filepath = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Feb'26 consolidated.xlsx")
    if not os.path.exists(filepath):
        return []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names
    except Exception:
        return []


@app.get("/api/consolidated-headers")
async def consolidated_headers():
    """Return per-sheet headers from the consolidated report."""
    filepath = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "Feb'26 consolidated.xlsx")
    result = {}
    if not os.path.exists(filepath):
        return result
    try:
        import openpyxl
        wb = openpyxl.load_workbook(filepath, read_only=True)
        for sname in wb.sheetnames:
            ws = wb[sname]
            row = next(ws.iter_rows(values_only=True), [])
            headers = [_normalize_header(str(c)) for c in row if c is not None and str(c).strip()]
            result[sname] = headers
        wb.close()
    except Exception:
        pass
    return result


@app.get("/api/consolidated-preview")
async def consolidated_preview(sheet_name: str = ""):
    """Return headers + sample rows from the consolidated report for a given sheet."""
    if not sheet_name:
        return {"headers": [], "rows": []}
    return get_consolidated_preview(sheet_name, max_rows=20)


@app.get("/api/consolidated-sheets")
async def consolidated_sheets():
    """List available sheets in the consolidated report for target selection."""
    sheets = list_consolidated_sheets()
    return {"sheets": sheets}


@app.get("/api/ai-status")
async def ai_status():
    """Check if any fallback API keys are configured in the server's .env file."""
    has_key = bool(
        os.getenv("GROQ_API_KEY") or 
        os.getenv("OPENROUTER_API_KEY") or 
        os.getenv("GEMINI_API_KEY")
    )
    return {"configured": has_key}



@app.post("/api/ai-auto-match")
async def ai_auto_match(data: dict):
    sheet_name = data.get("sheet_name", "").strip()
    filename = data.get("filename", "").strip()
    client_columns = data.get("columns", [])
    client_preview = data.get("preview", [])
    api_key = data.get("api_key", "").strip()

    # Fallback to .env keys if not passed by the client UI
    if not api_key:
        api_key = os.getenv("GROQ_API_KEY") or os.getenv("OPENROUTER_API_KEY") or os.getenv("GEMINI_API_KEY")
        if api_key:
            api_key = api_key.strip()

    if not api_key:
        raise HTTPException(400, "API Key is required. Please set it in .env or enter it in the UI.")

    # Determine sheets to align
    if sheet_name:
        all_target_sheets = [sheet_name]
    else:
        all_target_sheets = list_consolidated_sheets()
        if not all_target_sheets:
            all_target_sheets = ["Payment Tracker", "Master Data"]

    target_configs = {}
    for t_sheet in all_target_sheets:
        target_info = get_consolidated_preview(t_sheet, max_rows=5) # Latency optimization: limit target rows to 5
        target_headers = target_info["headers"]
        if not target_headers:
            if t_sheet == "Payment Tracker":
                target_headers = [
                    "S.no", "Sr No", "Client", "Assayer Name", "Assayer Code", "Assayer Phone",
                    "Assayer PAN", "Location", "State", "Zone", "Branch", "Branch Code",
                    "Month", "Audit Month & Year", "Type of Audit", "No. of Visits",
                    "Base Audit Fee", "Total pay (Base)", "Travel charges",
                    "Cancelled visits", "Branch Cancellation Charges",
                    "Andaman & Nicobar Branch Expenses", "Error Deduction",
                    "Total pay", "Remarks", "PAN Number", "Bank Name",
                    "A/c Number", "IFSC Code", "Schedule date", "Audit Status",
                    "Audit completion date", "No of days audited", "No of Packets audited",
                ]
            else:
                target_headers = [
                    "S.no", "Client", "Client fee", "Additional", "Final Client Fees", 
                    "Assayer fee", "Additional fee", "Distance", "Base Location", "Cancelled",
                    "Total", "Audit Remarks", "Contact Person", "SOL ID", "State", "Month"
                ]
        target_configs[t_sheet] = {
            "headers": target_headers,
            "preview": target_info["rows"][:5] # Latency optimization: limit rows to 5
        }

    # Build target sheets prompt description
    target_sheets_desc = ""
    for t_sheet, config in target_configs.items():
        target_sheets_desc += f"""
=== TARGET SHEET: "{t_sheet}" ===
Target column headers:
{json.dumps(config["headers"])}

Sample data rows from "{t_sheet}":
{json.dumps(config["preview"][:5], indent=2)}
"""

    client_rows_json = json.dumps(client_preview[:10], indent=2) # Latency optimization: limit source to 10 rows instead of 20

    # Build the prompt
    if len(all_target_sheets) == 1:
        # Single sheet mode for backward compatibility
        prompt = f"""You are an expert data migration and Excel consolidation architect.

We are building a mapping template that maps columns in a raw client Excel sheet to standard "target" database columns of a consolidated report.

=== CONTEXT ===
Active worksheet: "{all_target_sheets[0]}"
Uploaded filename: "{filename}"

=== TARGET (Standard Database) ===
Target column headers:
{json.dumps(target_configs[all_target_sheets[0]]["headers"])}

Sample data rows from target:
{json.dumps(target_configs[all_target_sheets[0]]["preview"][:5], indent=2)}

=== SOURCE (Uploaded Client Excel) ===
Source column headers from uploaded file:
{json.dumps(client_columns)}

Sample data rows from source:
{client_rows_json}

=== YOUR TASK ===
Analyze the source columns AND their sample data, then compare them against the target columns and target sample data. For each target column, determine:

1. **mappings**: Which source column (if any) should map to this target column? Leave as "" if no correspond.
2. **rules**: Determine datatype, mandatory, default_value, copy_from_column.

Return ONLY valid JSON with this exact structure (no markdown, no explanation):
{{
  "mappings": {{
    "Target Column 1": "Source Column 1"
  }},
  "rules": {{
    "Target Column 1": {{
      "datatype": "string | integer | decimal | date | time",
      "mandatory": true | false,
      "default_value": "suggested default or empty string",
      "copy_from_column": "Other target column or empty string"
    }}
  }}
}}
"""
    else:
        # High performance global multi-sheet mode
        prompt = f"""You are an expert data migration and Excel consolidation architect.

We are building a mapping template that maps columns in a raw client Excel sheet to standard "target" database columns across multiple worksheets of a consolidated report.

=== CONTEXT ===
Uploaded filename: "{filename}"

=== SOURCE (Uploaded Client Excel) ===
Source column headers:
{json.dumps(client_columns)}

Sample data rows from source:
{client_rows_json}

=== TARGET SCHEMAS (Standard Database Sheets) ===
{target_sheets_desc}

=== YOUR TASK ===
Analyze the source columns AND their sample data, then compare them against the target columns and target sample data for each target sheet. 
For EACH target sheet, determine mappings and rules (datatype, mandatory, default_value, copy_from_column).
Leave mapping as "" if a target column has no semantic source column.

Return ONLY valid JSON with this exact structure (no markdown, no explanation):
{{
  "sheets": {{
    "Payment Tracker": {{
      "mappings": {{
        "Target Column 1": "Source Column 1 or empty string"
      }},
      "rules": {{
        "Target Column 1": {{
          "datatype": "string | integer | decimal | date | time",
          "mandatory": true | false,
          "default_value": "suggested default or empty string",
          "copy_from_column": "Other target column or empty string"
        }}
      }}
    }},
    "Master Data": {{
      "mappings": {{
        "Target Column X": "Source Column X or empty string"
      }},
      "rules": {{
        "Target Column X": {{
          "datatype": "string | integer | decimal | date | time",
          "mandatory": true | false,
          "default_value": "suggested default or empty string",
          "copy_from_column": "Other target column or empty string"
        }}
      }}
    }}
  }}
}}
"""

    # Build prioritized queue of (provider, key) combinations to try
    queue = []
    
    # If the user passed a specific key in the frontend prompt, try that provider first
    if api_key:
        if api_key.startswith("gsk_"):
            queue.append(("groq", api_key))
        elif api_key.startswith("sk-or-"):
            queue.append(("openrouter", api_key))
        else:
            queue.append(("gemini", api_key))

    # Add environmental fallback keys from .env if not already in the queue
    env_gemini = os.getenv("GEMINI_API_KEY")
    env_groq = os.getenv("GROQ_API_KEY")
    env_openrouter = os.getenv("OPENROUTER_API_KEY")
    
    if env_gemini and env_gemini.strip() not in [k for _, k in queue]:
        queue.append(("gemini", env_gemini.strip()))
    if env_groq and env_groq.strip() not in [k for _, k in queue]:
        queue.append(("groq", env_groq.strip()))
    if env_openrouter and env_openrouter.strip() not in [k for _, k in queue]:
        queue.append(("openrouter", env_openrouter.strip()))

    if not queue:
        raise HTTPException(400, "API Key is required. Please set it in .env or enter it in the UI.")

    res_body = None
    resolved_provider = None
    error_list = []

    def _has_status(e: Exception, code: str) -> bool:
        return code in str(e)

    # Execute request with automatic fallback rotation
    for prov_idx, (prov, key) in enumerate(queue):
        if prov_idx > 0:
            print("  [Fallback] Rate limited on previous provider, waiting 30s before next...")
            time.sleep(30)

        retry_delay = 10
        print(f"[AI Auto-Match] Attempting template alignment using provider: {prov}...")
        
        if prov == "groq":
            url = "https://api.groq.com/openai/v1/chat/completions"
            headers = {
                "Authorization": f"Bearer {key}",
                "Content-Type": "application/json",
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            }
            for model in ["llama-3.3-70b-versatile", "gemma2-9b-it"]:
                payload = {
                    "model": model,
                    "messages": [{"role": "user", "content": prompt}],
                    "temperature": 0.1,
                }
                if model == "llama-3.3-70b-versatile":
                    payload["response_format"] = {"type": "json_object"}
                try:
                    req = urllib.request.Request(
                        url,
                        data=json.dumps(payload).encode("utf-8"),
                        headers=headers,
                        method="POST"
                    )
                    with urllib.request.urlopen(req, timeout=30) as response:
                        res_body = response.read().decode("utf-8")
                    resolved_provider = "groq"
                    break
                except Exception as e:
                    print(f"  [Groq] Model {model} failed: {str(e)}")
                    error_list.append(f"Groq ({model}): {str(e)}")
                    if _has_status(e, "429"):
                        time.sleep(retry_delay)
                        retry_delay = min(retry_delay + 10, 60)
            if res_body:
                break

        elif prov == "openrouter":
            url = "https://openrouter.ai/api/v1/chat/completions"
            headers = {
                "Authorization": f"Bearer {key}",
                "Content-Type": "application/json",
                "HTTP-Referer": "http://localhost:5173",
                "X-Title": "Consolidation Report Template Mapper",
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            }
            for model in ["google/gemma-4-31b-it:free", "meta-llama/llama-3.3-70b-instruct:free", "deepseek/deepseek-r1:free"]:
                payload = {
                    "model": model,
                    "messages": [{"role": "user", "content": prompt}],
                    "temperature": 0.1,
                    "response_format": {"type": "json_object"}
                }
                try:
                    req = urllib.request.Request(
                        url,
                        data=json.dumps(payload).encode("utf-8"),
                        headers=headers,
                        method="POST"
                    )
                    with urllib.request.urlopen(req, timeout=30) as response:
                        res_body = response.read().decode("utf-8")
                    resolved_provider = "openrouter"
                    break
                except Exception as e:
                    print(f"  [OpenRouter] Model {model} failed: {str(e)}")
                    error_list.append(f"OpenRouter ({model}): {str(e)}")
                    if _has_status(e, "429"):
                        time.sleep(retry_delay)
                        retry_delay = min(retry_delay + 10, 60)
            if res_body:
                break

        elif prov == "gemini":
            headers = {
                "Content-Type": "application/json",
                "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
            }
            for gemini_model in ["gemini-2.5-flash", "gemini-2.0-flash", "gemini-2.0-flash-lite"]:
                url = f"https://generativelanguage.googleapis.com/v1beta/models/{gemini_model}:generateContent?key={key}"
                payload = {
                    "contents": [{"parts": [{"text": prompt}]}],
                    "generationConfig": {"responseMimeType": "application/json"}
                }
                try:
                    req = urllib.request.Request(
                        url,
                        data=json.dumps(payload).encode("utf-8"),
                        headers=headers,
                        method="POST"
                    )
                    with urllib.request.urlopen(req, timeout=30) as response:
                        res_body = response.read().decode("utf-8")
                    resolved_provider = "gemini"
                    break
                except Exception as e:
                    print(f"  [Gemini] {gemini_model} failed: {str(e)}")
                    error_list.append(f"Gemini ({gemini_model}): {str(e)}")
                    if _has_status(e, "429"):
                        time.sleep(retry_delay)
                        retry_delay = min(retry_delay + 10, 60)
            if res_body:
                break

    if not res_body:
        advice = "Try: (1) wait 30-60 seconds and retry, (2) use a different API key, (3) add a Groq or OpenRouter API key in .env for more fallback options"
        raise HTTPException(502, f"AI Auto-Match failed: All available providers failed. {advice}. Errors: {' | '.join(error_list)}")

    try:
        res_json = json.loads(res_body)
        
        if resolved_provider in ("groq", "openrouter"):
            if "choices" not in res_json or not res_json["choices"]:
                raise HTTPException(502, f"LLM response has no choices: {res_json}")
            text_out = res_json["choices"][0]["message"]["content"].strip()
        else:
            if "candidates" not in res_json or not res_json["candidates"]:
                raise HTTPException(502, f"Gemini response has no candidates: {res_json}")
            text_out = res_json["candidates"][0]["content"]["parts"][0]["text"].strip()

        # Clean up markdown code block wrappers if present (e.g. from gemini-1.5-flash fallback)
        if text_out.startswith("```"):
            lines = text_out.split("\n")
            if lines[0].startswith("```"):
                lines = lines[1:]
            if lines and lines[-1].startswith("```"):
                lines = lines[:-1]
            text_out = "\n".join(lines).strip()

        mapping_result = json.loads(text_out)
        
        if len(all_target_sheets) == 1:
            return {
                "success": True,
                "mappings": mapping_result.get("mappings", {}),
                "rules": mapping_result.get("rules", {}),
                "target_headers": target_configs[all_target_sheets[0]]["headers"],
                "matched_sheet": all_target_sheets[0]
            }
        else:
            return {
                "success": True,
                "sheets": mapping_result.get("sheets", {}),
                "target_sheets": all_target_sheets
            }
    except Exception as e:
        raise HTTPException(502, f"AI Auto-Match failed to parse LLM response: {str(e)}")


@app.post("/api/schemas/from-mapping")
async def create_schema_from_mapping(data: dict):
    """Create a schema from a simplified drag-drop mapping result."""
    client_id = data.get("client_id", "").strip()
    if not client_id:
        raise HTTPException(400, "client_id is required")
    if _find_schema_file(client_id):
        raise HTTPException(409, f"Schema '{client_id}' already exists")

    sheets_in = data.get("sheets", [])
    if not sheets_in:
        raise HTTPException(400, "At least one sheet is required")

    sheets_out = {}
    for sh in sheets_in:
        sname = sh.get("name", "").strip()
        if not sname:
            continue
        col_defs = []
        for col in sh.get("columns", []):
            cname = col.get("canonical_name", "").strip()
            if not cname:
                continue
            entry = {"canonical_name": cname, "datatype": col.get("datatype", "string")}
            syns = [s.strip() for s in col.get("synonyms", []) if s.strip()]
            if syns:
                entry["synonyms"] = syns
            if col.get("mandatory"):
                entry["mandatory"] = True
            dv = col.get("default_value")
            if dv is not None and dv != "":
                try:
                    entry["default_value"] = int(dv) if "." not in str(dv) else float(dv)
                except (ValueError, TypeError):
                    entry["default_value"] = dv
            rx = col.get("validation_regex", "").strip()
            if rx:
                entry["validation_regex"] = rx
            excs = [e.strip() for e in col.get("validation_exceptions", []) if e.strip()]
            if excs:
                entry["validation_exceptions"] = excs
            cfc = col.get("copy_from_column", "").strip()
            if cfc:
                entry["copy_from_column"] = cfc
            hn = col.get("header_name", "").strip()
            if hn:
                entry["header_name"] = hn
            col_defs.append(entry)

        sheet_def = {
            "header_row": sh.get("header_row", 1),
            "data_start_row": sh.get("data_start_row", 2),
        }
        if sh.get("client_column"):
            sheet_def["client_column"] = sh["client_column"]
        if sh.get("s_no_column"):
            sheet_def["s_no_column"] = sh["s_no_column"]
        sheet_def["columns"] = col_defs

        sum_cols = [s.strip() for s in sh.get("sum_columns", []) if s.strip()]
        if sum_cols:
            sheet_def["sum_columns"] = sum_cols
        hidden = sh.get("hidden_columns", [])
        if hidden:
            sheet_def["hidden_columns"] = hidden
        order = sh.get("column_order", [])
        if order:
            sheet_def["column_order"] = order

        sheets_out[sname] = sheet_def

    schema = {
        "client_id": client_id,
        "client_display_name": data.get("client_display_name", "").strip() or client_id,
        "filename_pattern": data.get("filename_pattern", f"*{client_id}*"),
        "active": data.get("active", True),
        "sheets": sheets_out,
    }

    # Store rules as inline metadata if provided
    rules = data.get("rules", [])
    if rules:
        schema["rules"] = rules

    fp = os.path.join(SCHEMAS_DIR, f"{client_id}.yaml")
    os.makedirs(SCHEMAS_DIR, exist_ok=True)
    # Validate schema before writing — catches structural issues early
    try:
        SchemaDefinition.model_validate(schema)
    except Exception as e:
        raise HTTPException(422, f"Schema validation failed: {str(e)}")
    _write_schema_yaml(fp, schema)
    return {"success": True, "client_id": client_id, "filename": f"{client_id}.yaml"}


# Mount static files for the React frontend when built
if os.path.exists(frontend_dist):
    app.mount("/", StaticFiles(directory=frontend_dist, html=True), name="static")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run("src.web_api:app", host="0.0.0.0", port=8000, reload=True)

