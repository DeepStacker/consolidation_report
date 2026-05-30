import os
import openpyxl
import pandas as pd
from datetime import datetime, date
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# Canonical column headers for Consolidated Payment Tracker
PT_COLUMNS = [
    "S.no", "client", "Assayer Name", "Assayer Code", "Assayer Phone", "Location", "State", "Zone",
    "Audit Month & Year", "Type of Audit", "No. of Visits", "Base Audit Fee", "Total pay (Base)",
    " Travel charges(If any)", "Cancelled visits", "Branch Cancellation Charges",
    " Andaman & Nicobar Branch Expenses", "Error Deduction", "Total pay", "Remarks (if any)",
    "PAN Number", "Bank Name", "A/c Number", "IFSC Code"
]

# Canonical column headers for Consolidated Master Data
MD_COLUMNS = [
    "Sr No", "Client", "Month", "Zone", "SOL ID", "BRANCH", "Location ", "State", "Total No.of A/cs",
    "Assayer Name", "Assayer Code", "AssayerPhone", "Assayer PAN", "Contact Person", "Schedule date",
    "Audit \nStatus", "Audit \ncompletion date", "No of days \naudited ", "No of days \naudited For client",
    "No of Packets \naudited", "Additional Packet", "Assayer Reporting time at Branch", "Audit start time",
    "Audit End Time", "Client fee", "Additional", "Final Client Fees", "Assayer fee", "Additional fee",
    "Distance", "Base Location", "Remarks", "Assayer fee.1", "Additional fee.1", "Cancelled",
    "Error Deduciton", "Total", "Audit Remarks", "Seeding Status", "Report",
    "Total pouches suggested for audit", "Already Audited", "A/C Closed", "A/C Auctioned", "Packet Missing",
    "Actual Audited (except already audited & A/C closed)  ", "Extra audited pouches",
    "Total No.of packets actually audited", "Type oof Audit", "Aadhar No", "Remarks.1", "Urban/ Rural",
    "T & F Client fee", "POA Client Fees", "Cancelled\nClient fee", "billing  remarks",
    "Touch and Feel Audit Packet Count", "POA \nPacket count", "Additional Packet  T& F",
    "Additional Packet  POA", "oracle_id", "Address"
]

def format_cell(cell, font_name="Calibri", font_size=11, bold=False, italic=False, color="000000",
                alignment=None, fill=None, border=None, number_format=None):
    """Utility to format an openpyxl cell cleanly."""
    cell.font = Font(name=font_name, size=font_size, bold=bold, italic=italic, color=color)
    if alignment:
        cell.alignment = alignment
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format

def write_consolidated_workbook(axis_pt, rbl_pt, axis_md, rbl_md, output_path):
    """Generates the clean Consolidated Excel Workbook with dynamic formula ranges."""
    wb = openpyxl.Workbook()
    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # -------------------------------------------------------------------------
    # SHEET 1: Payment Tracker
    # -------------------------------------------------------------------------
    ws_pt = wb.create_sheet(title="Payment Tracker")
    
    # 1. Write headers
    header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Sleek Navy Blue
    header_font_color = "FFFFFF"
    
    for c_idx, col_name in enumerate(PT_COLUMNS, 1):
        cell = ws_pt.cell(row=1, column=c_idx, value=col_name)
        format_cell(
            cell, bold=True, color=header_font_color, fill=header_fill,
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
        )
    ws_pt.row_dimensions[1].height = 28

    # 2. Write Data Rows
    # Combine Axis and RBL Payment Tracker rows
    pt_rows = []
    
    # Axis rows
    if not axis_pt.empty:
        # Sort or maintain order
        for _, row in axis_pt.iterrows():
            pt_rows.append((row, "Axis Bank POA"))
            
    # RBL rows
    if not rbl_pt.empty:
        for _, row in rbl_pt.iterrows():
            pt_rows.append((row, "RBL(muthoot)"))

    # Populate cells
    s_no = 1
    current_row = 2
    
    for row_data, client_name in pt_rows:
        # Set Sequential S.no
        ws_pt.cell(row=current_row, column=1, value=s_no)
        ws_pt.cell(row=current_row, column=2, value=client_name)
        
        for c_idx in range(3, len(PT_COLUMNS) + 1):
            col_name = PT_COLUMNS[c_idx - 1]
            val = row_data.get(col_name, None)
            
            # Format Excel formulas or standard cell values
            cell = ws_pt.cell(row=current_row, column=c_idx, value=val)
            
            # Formatter defaults: align numbers right, text left
            if isinstance(val, (int, float)):
                cell.number_format = "#,##0.00" if isinstance(val, float) else "0"
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(horizontal="left")
                
        ws_pt.row_dimensions[current_row].height = 20
        s_no += 1
        current_row += 1

    # 3. Dynamic Sum coordinates injection (circuit breaker range overflow solution)
    data_end_row = current_row - 1
    total_row = data_end_row + 2  # Leaves one row empty (row N+1), puts totals at row N+2
    
    # Set the total label in column A (1)
    cell_tot_label = ws_pt.cell(row=total_row, column=1, value="Total")
    format_cell(cell_tot_label, bold=True, alignment=Alignment(horizontal="center"))
    
    # Calculate column indexes for formulas
    # Column M (13) is 'Total pay (Base)'
    # Column S (19) is 'Total pay'
    cell_base_total = ws_pt.cell(row=total_row, column=13, value=f"=SUM(M2:M{data_end_row})")
    format_cell(cell_base_total, bold=True, number_format="#,##0.00", alignment=Alignment(horizontal="right"))
    
    cell_grand_total = ws_pt.cell(row=total_row, column=19, value=f"=SUM(S2:S{data_end_row})")
    format_cell(cell_grand_total, bold=True, number_format="#,##0.00", alignment=Alignment(horizontal="right"))
    
    ws_pt.row_dimensions[total_row].height = 22

    # Border style for total row
    double_bottom_border = Border(
        top=Side(style='thin', color='000000'),
        bottom=Side(style='double', color='000000')
    )
    for c_idx in range(1, len(PT_COLUMNS) + 1):
        ws_pt.cell(row=total_row, column=c_idx).border = double_bottom_border

    # Auto-adjust column widths
    for col in ws_pt.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_pt.column_dimensions[col_letter].width = max(max_len + 3, 12)


    # -------------------------------------------------------------------------
    # SHEET 2: Master Data
    # -------------------------------------------------------------------------
    ws_md = wb.create_sheet(title="Master Data")
    
    # 1. Write headers
    md_header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # Steel Blue
    
    for c_idx, col_name in enumerate(MD_COLUMNS, 1):
        cell = ws_md.cell(row=1, column=c_idx, value=col_name)
        format_cell(
            cell, bold=True, color=header_font_color, fill=md_header_fill,
            alignment=Alignment(horizontal="center", vertical="center", wrap_text=True)
        )
    ws_md.row_dimensions[1].height = 32

    # 2. Write Data Rows
    # Combine Axis and RBL (which has POA and GL) DataFrames
    md_rows = []
    if not axis_md.empty:
        for _, row in axis_md.iterrows():
            md_rows.append((row, "Axis Bank POA"))
    if not rbl_md.empty:
        for _, row in rbl_md.iterrows():
            # RBL rows can be POA or GL (Gold Loan).
            # Determine client name from row values
            client_tag = row.get("Client", "RBL(muthoot)")
            md_rows.append((row, client_tag))

    md_s_no = 1
    current_md_row = 2
    
    for row_data, client_name in md_rows:
        # Sequential re-numbering
        ws_md.cell(row=current_md_row, column=1, value=float(md_s_no))
        ws_md.cell(row=current_md_row, column=2, value=client_name)
        
        for c_idx in range(3, len(MD_COLUMNS) + 1):
            col_name = MD_COLUMNS[c_idx - 1]
            val = row_data.get(col_name, None)
            
            # Format datetime values cleanly
            if isinstance(val, (datetime, date)) and not pd.isna(val):
                cell = ws_md.cell(row=current_md_row, column=c_idx, value=val.strftime("%Y-%m-%d"))
                cell.alignment = Alignment(horizontal="center")
            elif isinstance(val, (datetime, date)) and pd.isna(val):
                cell = ws_md.cell(row=current_md_row, column=c_idx, value=None)
            else:
                cell = ws_md.cell(row=current_md_row, column=c_idx, value=val)
                if isinstance(val, (int, float)):
                    if not pd.isna(val):
                        # Use integer format for counts, decimal for fees
                        cell.number_format = "#,##0.00" if "fee" in col_name.lower() or "pay" in col_name.lower() or col_name in ["Cancelled", "Total", "Additional"] else "0"
                        cell.alignment = Alignment(horizontal="right")
                else:
                    cell.alignment = Alignment(horizontal="left")
                    
        ws_md.row_dimensions[current_md_row].height = 18
        md_s_no += 1
        current_md_row += 1

    # Auto-adjust column widths
    for col in ws_md.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws_md.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 30)

    # 3. Apply Column Hiding (Columns 10 to 15, 0-indexed: Column J to O)
    # 10: Assayer Name (J), 11: Assayer Code (K), 12: AssayerPhone (L), 13: Assayer PAN (M), 14: Contact Person (N), 15: Schedule date (O)
    for c_idx in range(10, 16):
        col_letter = get_column_letter(c_idx)
        ws_md.column_dimensions[col_letter].hidden = True

    # Save completed workbook
    wb.save(output_path)
    print(f"Consolidated workbook written successfully to: {output_path}")
    return output_path
