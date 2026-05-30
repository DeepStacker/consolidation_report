import os
import glob
import yaml
import openpyxl
import pandas as pd
import numpy as np
from datetime import datetime, date, time

def load_yaml_schema(schema_path):
    """Loads and returns a YAML schema file."""
    with open(schema_path, 'r') as f:
        return yaml.safe_load(f)

def clean_value(val, datatype, default_val=None, col_name=""):
    """Casts raw cell values to canonical Python types based on datatype."""
    if val is None or (isinstance(val, str) and val.strip() == "") or (isinstance(val, float) and np.isnan(val)):
        return default_val

    if datatype == "integer":
        try:
            # Strip decimal points if integer was parsed as float (e.g. 1.0 -> 1)
            if isinstance(val, float) or isinstance(val, str):
                return int(float(val))
            return int(val)
        except Exception:
            return default_val if default_val is not None else 0

    elif datatype == "decimal":
        try:
            return float(val)
        except Exception:
            return default_val if default_val is not None else 0.0

    elif datatype == "date":
        if isinstance(val, datetime):
            return val
        if isinstance(val, date):
            return datetime.combine(val, time.min)
        if isinstance(val, str):
            for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%Y/%m/%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
                try:
                    return datetime.strptime(val.strip(), fmt)
                except ValueError:
                    continue
        return default_val

    elif datatype == "time":
        if isinstance(val, time):
            return val
        if isinstance(val, datetime):
            return val.time()
        return str(val).strip()

    elif datatype == "string":
        # Handle float values that represent integers formatted as strings (e.g. SOL ID 55.0 -> 55)
        if isinstance(val, float):
            if val.is_integer():
                return str(int(val))
            return str(val)
        return str(val).strip()

    return val

def ingest_sheet(file_path, sheet_name, sheet_config):
    """Ingests a sheet based on sheet schema configurations and returns a clean list of dictionaries."""
    wb = openpyxl.load_workbook(file_path, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook '{file_path}'")

    ws = wb[sheet_name]
    header_row_idx = sheet_config.get("header_row", 1)
    data_start_row_idx = sheet_config.get("data_start_row", 2)
    columns_config = sheet_config.get("columns", [])

    # Extract raw headers
    max_col = ws.max_column
    raw_headers = []
    for c in range(1, max_col + 1):
        header_val = ws.cell(row=header_row_idx, column=c).value
        raw_headers.append(header_val if header_val is not None else "")

    # Build synonym mapping
    col_mapping = {}  # col_idx (1-indexed) -> canonical_column_metadata
    for col_meta in columns_config:
        canonical_name = col_meta["canonical_name"]
        synonyms = [s.strip().lower() for s in col_meta.get("synonyms", [])]
        
        # Add canonical name itself as a synonym
        synonyms.append(canonical_name.strip().lower())
        
        # Check sheet headers
        for idx, h in enumerate(raw_headers):
            h_clean = str(h).strip().lower()
            if h_clean in synonyms:
                col_mapping[idx + 1] = col_meta
                break

    # Read data rows
    max_row = ws.max_row
    clean_rows = []
    
    for r in range(data_start_row_idx, max_row + 1):
        s_no_val = ws.cell(row=r, column=1).value
        
        # Rule Check: Exclude row if the first 3 columns are completely empty, 
        # or if the first column contains the word "Total" (case-insensitive).
        # This preserves active data rows that have missing sequence numbers (e.g. Guruprasad R Shet).
        first_three = [ws.cell(row=r, column=c).value for c in range(1, min(4, max_col + 1))]
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in first_three):
            break
            
        if s_no_val is not None and isinstance(s_no_val, str) and "total" in s_no_val.lower():
            break
            
        # Also exclude completely empty rows
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, max_col + 1)]
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row_vals):
            continue
            
        # Parse data cells based on canonical mapping
        row_dict = {}
        for c in range(1, max_col + 1):
            if c in col_mapping:
                col_meta = col_mapping[c]
                canonical_name = col_meta["canonical_name"]
                datatype = col_meta["datatype"]
                default_val = col_meta.get("default_value", None)
                
                cell_val = ws.cell(row=r, column=c).value
                cleaned_val = clean_value(cell_val, datatype, default_val, canonical_name)
                row_dict[canonical_name] = cleaned_val
        
        # Seed default values for mapped columns not present in the sheet
        for col_meta in columns_config:
            canonical_name = col_meta["canonical_name"]
            if canonical_name not in row_dict:
                row_dict[canonical_name] = col_meta.get("default_value", None)
                
        clean_rows.append(row_dict)

    return clean_rows

def scan_and_ingest_client(workspace_path, schema_path):
    """Scans the workspace for files matching the client schema, and ingests all configured sheets."""
    schema = load_yaml_schema(schema_path)
    client_id = schema["client_id"]
    filename_pattern = schema["filename_pattern"]
    sheets_config = schema["sheets"]
    
    # Search for client file
    search_pattern = os.path.join(workspace_path, filename_pattern)
    matched_files = glob.glob(search_pattern)
    if not matched_files:
        # Try case-insensitive matching if first match fails
        all_files = glob.glob(os.path.join(workspace_path, "*"))
        matched_files = [f for f in all_files if filename_pattern.replace("*", "").lower() in os.path.basename(f).lower()]
        
    if not matched_files:
        raise FileNotFoundError(f"No file found matching pattern '{filename_pattern}' for client '{client_id}' in workspace '{workspace_path}'")
        
    # Take the first matched file
    client_file = matched_files[0]
    client_data = {}
    
    for sheet_name, sheet_config in sheets_config.items():
        clean_rows = ingest_sheet(client_file, sheet_name, sheet_config)
        client_data[sheet_name] = pd.DataFrame(clean_rows)
        
    return client_id, client_file, client_data
