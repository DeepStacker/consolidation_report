import pytest
import os
import openpyxl
import pandas as pd
from typing import Generator

@pytest.fixture
def test_workspace(tmp_path) -> str:
    """Fixture that initializes a temporary test workspace directory."""
    workspace_dir = os.path.join(tmp_path, "consolidation_report")
    os.makedirs(workspace_dir, exist_ok=True)
    os.makedirs(os.path.join(workspace_dir, "config", "schemas"), exist_ok=True)
    return workspace_dir

@pytest.fixture
def mock_axis_tracker(test_workspace) -> str:
    """Fixture that generates a mock Axis Bank POA Excel workbook containing clean rows and spacers."""
    filepath = os.path.join(test_workspace, "Axis Bank POA Payment Tracker - Feb'26.xlsx")
    wb = openpyxl.Workbook()
    
    # 1. Sheet 1: Payment Tracker
    ws_pt = wb.active
    ws_pt.title = "Payment Tracker"
    
    headers_pt = [
        "S.no", "Assayer Name", "Assayer Code", "Assayer Phone", "Location", "State", "Zone",
        "Audit Month & Year", "Type of Audit", "No. of Visits", "Base Audit Fee", "Total pay (Base)",
        " Travel charges(If any)", "Cancelled visits", "Branch Cancellation Charges",
        " Andaman & Nicobar Branch Expenses", "Error Deduction", "Total pay", "Remarks (if any)",
        "PAN Number", "Bank Name", "A/c Number", "IFSC Code"
    ]
    for c, h in enumerate(headers_pt, 1):
        ws_pt.cell(row=1, column=c, value=h)
        
    # Active data row
    row_pt = [
        1.0, "Krishna Gopal Verma", "AD0351", 9616953240, "Menhdawal", "Uttar Pradesh", "North",
        "Feb'26", "AXIS Bank POA", 8.0, 1000.0, 8000.0, 5500.0, 0.0, 0.0, 0.0, 0.0, 13500.0,
        "Travel Fee added", "AUDPV5880C", "Indian Bank", 5029236450, "IDIB000M692"
    ]
    for c, val in enumerate(row_pt, 1):
        ws_pt.cell(row=2, column=c, value=val)
        
    # Totals row at bottom
    ws_pt.cell(row=3, column=1, value="Total")
    ws_pt.cell(row=3, column=13, value="=SUM(M2:M2)")
    ws_pt.cell(row=3, column=19, value="=SUM(S2:S2)")

    # 2. Sheet 2: Master Data
    ws_md = wb.create_sheet(title="Master Data")
    headers_md = [
        "Sr No", "Month", "Zone", "SOL ID", "BRANCH", "State", "Assayer Name", "Assayer Code",
        "AssayerPhone", "Assayer PAN", "Contact Person", "Schedule date", "Audit Status",
        "Audit completion date", "No of days audited ", "No of days audited For client",
        "No of Packets audited", "Additional Packet", "Assayer Reporting time at Branch",
        "Audit start time", "Audit End Time", "Client fee", "Additional", "Final Client Fees",
        "Assayer fee", "Additional fee", "Distance", "Base Location", "Remarks", "Assayer fee.1",
        "Additional fee.1", "Cancelled", "Error Deduciton", "Total", "Audit Remarks"
    ]
    for c, h in enumerate(headers_md, 1):
        ws_md.cell(row=1, column=c, value=h)
        
    # Active data row
    row_md = [
        1.0, "Feb'26", "North", "55", "Swasthya Vihar", "Delhi", "Ramesh Chand Verma", "AS0085",
        9810982340, "AFQPV2026J", "Parase", "2026-01-28", "Completed", "2026-02-02", 4.0, 1.0,
        215.0, 165.0, "10:00AM", "10:30AM", "5:30PM", 3000.0, 11550.0, 14550.0, 1000.0, 500.0,
        "100KM", "Delhi", "Travelling required", 4000.0, 2000.0, 0.0, 0.0, 6000.0, None
    ]
    for c, val in enumerate(row_md, 1):
        ws_md.cell(row=2, column=c, value=val)
        
    wb.save(filepath)
    return filepath

@pytest.fixture
def mock_rbl_tracker(test_workspace) -> str:
    """Fixture that generates a mock RBL Muthoot POA Excel workbook containing clean rows and spacer."""
    filepath = os.path.join(test_workspace, "RBL(Muthoot Fincorp) Payment Tracker -Feb'26.xlsx")
    wb = openpyxl.Workbook()
    
    # 1. Sheet 1: Payment Tracker
    ws_pt = wb.active
    ws_pt.title = "Payment Tracker"
    headers_pt = [
        "S.no", "Assayer Name", "Assayer Code", "Assayer Phone", "Location", "State", "Zone",
        "Audit Month & Year", "Type of Audit", "No of visits", "Base Audit Fee", "Total pay (Base)",
        " Travel charges(If any)", "Audit Cancellation Fees", "Total pay", "Remarks (if any)",
        "PAN Number", "Bank Name", "A/c Number", "IFSC Code"
    ]
    for c, h in enumerate(headers_pt, 1):
        ws_pt.cell(row=1, column=c, value=h)
        
    # Active data row
    row_pt = [
        1.0, "Rajeshwer Kumar", "AS0001", 9413933240, "Hanumangarh", "Rajasthan", "North",
        "Feb'26", "RBL Muthoot Fincorp POA", 6, 1000.0, 6000, 5500, 0, 11500,
        "Payment as per approval", "AFCPS0033A", "IDBI Bank", 373104000018124, "IBKL0000373"
    ]
    for c, val in enumerate(row_pt, 1):
        ws_pt.cell(row=2, column=c, value=val)
        
    # Totals row at bottom (S.no is None, totals at other cols)
    ws_pt.cell(row=3, column=1, value=None)
    ws_pt.cell(row=3, column=10, value=6)
    ws_pt.cell(row=3, column=12, value=6000)
    ws_pt.cell(row=3, column=15, value=11500)

    # 2. Sheet 2: Master Data
    ws_md = wb.create_sheet(title="Master Data")
    headers_md = [
        "Sr No.", "Zone", "State", "Branch Code", "BranchName", "Assayer Name", "Assayer Code",
        "AssayerPhone", "Assayer PAN", "Process Manager", "Month", "Audit schedule date", "Audit Status",
        "End Date", "No. of Visit", "Client Fees", "Final Client Fees", "Assayer Fees", "Additional Fees",
        "Assayer Base location", "Distance", "Remarks", "Total pouches suggested for audit", "A/C Closed",
        "A/C Auctioned", "Packet Missing", "Wrongly Created/Same account TR and Regular PKT",
        "Actual Audited (except already audited & A/C closed)", "Extra audited pouches",
        "Total No.of packets actually audited", "Assayer Fees.1", "Additional Fees.1", "Cancellation",
        "Cancelled Visit", "Total", "Other Remarks"
    ]
    for c, h in enumerate(headers_md, 1):
        ws_md.cell(row=1, column=c, value=h)
        
    # Active data row
    row_md = [
        1.0, "West", "MADHYA PRADESH", "F2733", "KOLAR-BHOPAL", "Nihal Soni", "AS0469",
        9074348240, "FVEPS3805P", "Krushna", "Feb'26", "2026-02-06", "Completed", "2026-02-06",
        1.0, 3000.0, 3000.0, 1000.0, 1000.0, "Katni", "650km", "Travelling Required", 7.0, 0.0,
        0.0, 0.0, 0.0, 0.0, 7.0, 7.0, 1000.0, 1000.0, 0.0, None, 2000.0, None
    ]
    for c, val in enumerate(row_md, 1):
        ws_md.cell(row=2, column=c, value=val)
        
    wb.save(filepath)
    return filepath
