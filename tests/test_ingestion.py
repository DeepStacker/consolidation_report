import pytest
import os
import openpyxl
from src.readers.excel_reader import (
    ingest_raw_rows, 
    compile_merged_cell_cache, 
    get_cell_value_cached,
    is_totals_row,
    recalculate_formulas_fallback
)
from src.schema_loader import load_schema_config
from src.models.exceptions import SchemaConfigException, MissingSheetException, IngestionException

def test_axis_sheet_ingestion(test_workspace, mock_axis_tracker):
    """Tests if Axis Payment Tracker workbook is ingested cleanly, ignoring spacer and totals row."""
    # Write mock axis poa yaml config schema into config/schemas
    axis_schema_yaml = """
client_id: "axis_poa"
client_display_name: "Axis Bank POA"
filename_pattern: "*Axis Bank POA*"
active: true
sheets:
  "Payment Tracker":
    header_row: 1
    data_start_row: 2
    columns:
      - canonical_name: "S.no"
        synonyms: ["S.no"]
        datatype: "integer"
        mandatory: true
      - canonical_name: "client"
        datatype: "string"
        default_value: "Axis Bank POA"
      - canonical_name: "Assayer Name"
        synonyms: ["Assayer Name"]
        datatype: "string"
      - canonical_name: "Assayer Code"
        synonyms: ["Assayer Code"]
        datatype: "string"
        validation_regex: "^AS[0-9]{4}$|^AD[0-9]{4}$"
        mandatory: true
      - canonical_name: "Total pay"
        synonyms: ["Total pay"]
        datatype: "decimal"
        mandatory: true
"""
    axis_config_path = os.path.join(test_workspace, "config", "schemas", "axis_poa.yaml")
    with open(axis_config_path, "w") as f:
        f.write(axis_schema_yaml)

    # 1. Ingest Raw Rows
    raw_rows = ingest_raw_rows(mock_axis_tracker, "Payment Tracker", 1, 2)
    # The raw sheet has 3 rows: row 1 is header, row 2 is clean data, row 3 is total.
    # The boundary checker should filter out Row 3, returning exactly 1 clean row!
    assert len(raw_rows) == 1
    assert raw_rows[0]["Assayer Name"] == "Krishna Gopal Verma"
    assert raw_rows[0]["Assayer Code"] == "AD0351"

    # 2. Schema Loader successfully parses config
    schema = load_schema_config(axis_config_path)
    assert schema.client_id == "axis_poa"
    assert schema.client_display_name == "Axis Bank POA"
    assert "Payment Tracker" in schema.sheets

def test_missing_sheet_throws_exception(test_workspace, mock_axis_tracker):
    """Tests that ingesting a non-existent sheet raises a MissingSheetException."""
    with pytest.raises(MissingSheetException):
        ingest_raw_rows(mock_axis_tracker, "Non Existent Sheet", 1, 2)

def test_corrupted_file_throws_exception(test_workspace):
    """Tests that ingesting a corrupted file triggers an IngestionException."""
    bad_filepath = os.path.join(test_workspace, "corrupted_file.xlsx")
    with open(bad_filepath, "w") as f:
        f.write("Definitely Not An Excel File Zip Package")
        
    with pytest.raises(IngestionException):
        ingest_raw_rows(bad_filepath, "Payment Tracker", 1, 2)

def test_schema_loader_missing_config():
    """Tests that loading a non-existent configuration raises SchemaConfigException."""
    with pytest.raises(SchemaConfigException):
        load_schema_config("non_existent_config_file.yaml")

def test_schema_loader_invalid_yaml(test_workspace):
    """Tests that loading an invalid YAML file raises SchemaConfigException."""
    bad_yaml_path = os.path.join(test_workspace, "bad_syntax.yaml")
    with open(bad_yaml_path, "w") as f:
        f.write("invalid: : yaml: syntax")
        
    with pytest.raises(SchemaConfigException):
        load_schema_config(bad_yaml_path)

def test_schema_loader_validation_failure(test_workspace):
    """Tests that loading a YAML with missing mandatory schema fields triggers SchemaConfigException."""
    invalid_schema_yaml = """
client_display_name: "Axis Bank POA"
# missing client_id and filename_pattern
active: true
"""
    invalid_config_path = os.path.join(test_workspace, "invalid_config.yaml")
    with open(invalid_config_path, "w") as f:
        f.write(invalid_schema_yaml)
        
    with pytest.raises(SchemaConfigException):
        load_schema_config(invalid_config_path)

def test_coordinate_grid_cache_lookup(test_workspace):
    """Tests that the pre-compiled Coordinate lookup cache correctly resolves merged cell locations."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Sheet"
    
    # Create a merged cell range and write to top-left
    ws.cell(row=2, column=2, value="Merged Block")
    ws.merge_cells(start_row=2, start_column=2, end_row=4, end_column=4)
    
    # Compile cache
    cache = compile_merged_cell_cache(ws)
    
    # Assert coordinates inside range map to top-left value
    assert get_cell_value_cached(ws, 2, 2, cache) == "Merged Block"
    assert get_cell_value_cached(ws, 3, 3, cache) == "Merged Block"
    assert get_cell_value_cached(ws, 4, 4, cache) == "Merged Block"
    
    # Assert coordinates outside range do not get cached
    assert get_cell_value_cached(ws, 1, 1, cache) is None
    assert get_cell_value_cached(ws, 5, 5, cache) is None
    
    wb.close()

def test_spacer_rows_in_middle_of_data(test_workspace):
    """P0 Test: Verifies that spacer rows placed in the middle of dataset are skipped, without halting ingestion early."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Tracker"
    
    # Headers in Row 1
    ws.cell(row=1, column=1, value="S.no")
    ws.cell(row=1, column=2, value="Assayer Name")
    ws.cell(row=1, column=3, value="Assayer Code")
    ws.cell(row=1, column=4, value="Total pay")
    
    # Row 2: Valid Row 1
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value="Krishna Gopal Verma")
    ws.cell(row=2, column=3, value="AD0351")
    ws.cell(row=2, column=4, value=13500.0)
    
    # Row 3: Blank Spacer Row in the middle of data! (Naively halt trigger)
    for c in range(1, 5):
        ws.cell(row=3, column=c, value=None)
        
    # Row 4: Valid Row 2 (Tests that this row is NOT lost)
    ws.cell(row=4, column=1, value=2)
    ws.cell(row=4, column=2, value="Rajeshwer Kumar")
    ws.cell(row=4, column=3, value="AS0001")
    ws.cell(row=4, column=4, value=10000.0)
    
    filepath = os.path.join(test_workspace, "test_spacer_mid.xlsx")
    wb.save(filepath)
    wb.close()
    
    # Ingest rows
    rows = ingest_raw_rows(filepath, "Payment Tracker", 1, 2)
    
    # Expected: The blank Row 3 is cleanly filtered, but the active Row 4 is successfully ingested!
    # Total data rows ingested must be exactly 2!
    assert len(rows) == 2
    assert rows[0]["Assayer Code"] == "AD0351"
    assert rows[1]["Assayer Code"] == "AS0001"

def test_totals_detection_variants():
    """P0 Test: Verifies totals row detector can identify totals blocks without catching legitimate essayers."""
    headers = ["S.no", "Assayer Name", "Assayer Code", "Total pay"]
    
    # Case A: Real transaction row (contains totals value in total pay column, but has name and code)
    legit_row = [1, "Guruprasad R Shet", "AS0701", 1000.0]
    assert is_totals_row(legit_row, headers) is False
    
    # Case B: True Totals Row (contains no assayer code or name, but has values in payment columns)
    totals_row = [None, None, None, 1577600.0]
    assert is_totals_row(totals_row, headers) is True
    
    # Case C: Real totals row with a word "Total" in first cell
    totals_word_row = ["Grand Total", None, None, 1577600.0]
    assert is_totals_row(totals_word_row, headers) is True

def test_duplicate_header_protection(test_workspace):
    """P0 Test: Verifies that duplicate column headers are renamed with suffix indices, preventing silent overwrite."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Tracker"
    
    # Headers in Row 1: duplicate 'Date' columns
    ws.cell(row=1, column=1, value="S.no")
    ws.cell(row=1, column=2, value="Date")  # first date
    ws.cell(row=1, column=3, value="Date")  # second date (duplicate!)
    ws.cell(row=1, column=4, value="Total pay")
    
    # Row 2: values
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value="2026-02-10")
    ws.cell(row=2, column=3, value="2026-02-11")
    ws.cell(row=2, column=4, value=5000.0)
    
    filepath = os.path.join(test_workspace, "test_dup_headers.xlsx")
    wb.save(filepath)
    wb.close()
    
    rows = ingest_raw_rows(filepath, "Payment Tracker", 1, 2)
    assert len(rows) == 1
    
    # Expected: The duplicate 'Date' columns must be suffixed to 'Date' and 'Date_1'
    assert "Date" in rows[0]
    assert "Date_1" in rows[0]
    assert rows[0]["Date"] == "2026-02-10"
    assert rows[0]["Date_1"] == "2026-02-11"

def test_scientific_notation_protection(test_workspace):
    """P0 Test: Verifies that account numbers read as floats are restored to exact integers, preventing digit loss."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payment Tracker"
    
    # Headers
    ws.cell(row=1, column=1, value="S.no")
    ws.cell(row=1, column=2, value="A/c Number")
    
    # Row 2: precise account number represented as float
    ws.cell(row=2, column=1, value=1)
    ws.cell(row=2, column=2, value=284801000003273.0)
    
    filepath = os.path.join(test_workspace, "test_scientific.xlsx")
    wb.save(filepath)
    wb.close()
    
    rows = ingest_raw_rows(filepath, "Payment Tracker", 1, 2)
    assert len(rows) == 1
    
    # Expected: Value is successfully cast as an exact integer
    assert isinstance(rows[0]["A/c Number"], int)
    assert rows[0]["A/c Number"] == 284801000003273

def test_formula_safety_fallback_layer():
    """P0 Test: Verifies that if formula columns return None/0.0, the safety layer recalculates them correctly."""
    uncalculated_row = {
        "No. of Visits": 8,
        "Base Audit Fee": 1000.0,
        "Total pay (Base)": None,  # uncalculated formula cell (None)
        " Travel charges(If any)": 5000.0,
        "Branch Cancellation Charges": 0.0,
        " Andaman & Nicobar Branch Expenses": 0.0,
        "Error Deduction": 1000.0,
        "Total pay": 0.0           # uncalculated formula net (0.0)
    }
    
    recalculated = recalculate_formulas_fallback(uncalculated_row)
    
    # Expected:
    # 1. Total pay (Base) = 8 * 1000.0 = 8000.0
    # 2. Total pay = 8000.0 (Base) + 5000.0 (Travel) - 1000.0 (Error) = 12000.0
    assert recalculated["Total pay (Base)"] == 8000.0
    assert recalculated["Total pay"] == 12000.0
