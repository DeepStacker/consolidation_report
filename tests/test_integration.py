import pytest
import os
import pandas as pd
from src.main import execute_e2e_consolidation

def test_pipeline_e2e_execution(test_workspace, mock_axis_tracker, mock_rbl_tracker):
    """Tests E2E pipeline execution with mock banking workbooks in temporary workspace."""
    # 1. Write the YAML configurations to config/schemas
    axis_yaml = """
client_id: "axis_poa"
client_display_name: "Axis Bank POA"
filename_pattern: "*Axis Bank POA*"
active: true
sheets:
  "Payment Tracker":
    header_row: 1
    data_start_row: 2
    client_column: "client"
    s_no_column: "S.no"
    sum_columns:
      - "Total pay (Base)"
      - "Total pay"
    columns:
      - canonical_name: "S.no"
        synonyms: ["S.no"]
        datatype: "integer"
        mandatory: true
      - canonical_name: "client"
        datatype: "string"
        default_value: "Axis Bank POA"
      - canonical_name: "Assayer Name"
        synonyms: ["Assayer Name"]
        datatype: "string"
      - canonical_name: "Assayer Code"
        synonyms: ["Assayer Code"]
        datatype: "string"
        mandatory: true
      - canonical_name: "Assayer Phone"
        synonyms: ["Assayer Phone"]
        datatype: "string"
      - canonical_name: "Location"
        synonyms: ["Location"]
        datatype: "string"
      - canonical_name: "State"
        synonyms: ["State"]
        datatype: "string"
      - canonical_name: "Zone"
        synonyms: ["Zone"]
        datatype: "string"
      - canonical_name: "Audit Month & Year"
        synonyms: ["Audit Month & Year"]
        datatype: "string"
      - canonical_name: "Type of Audit"
        synonyms: ["Type of Audit"]
        datatype: "string"
      - canonical_name: "No. of Visits"
        synonyms: ["No. of Visits"]
        datatype: "integer"
      - canonical_name: "Base Audit Fee"
        synonyms: ["Base Audit Fee"]
        datatype: "decimal"
      - canonical_name: "Total pay (Base)"
        synonyms: ["Total pay (Base)"]
        datatype: "decimal"
      - canonical_name: " Travel charges(If any)"
        synonyms: [" Travel charges(If any)"]
        datatype: "decimal"
      - canonical_name: "Cancelled visits"
        synonyms: ["Cancelled visits"]
        datatype: "integer"
      - canonical_name: "Branch Cancellation Charges"
        synonyms: ["Branch Cancellation Charges"]
        datatype: "decimal"
      - canonical_name: " Andaman & Nicobar Branch Expenses"
        synonyms: [" Andaman & Nicobar Branch Expenses"]
        datatype: "decimal"
      - canonical_name: "Error Deduction"
        synonyms: ["Error Deduction"]
        datatype: "decimal"
      - canonical_name: "Total pay"
        synonyms: ["Total pay"]
        datatype: "decimal"
        mandatory: true
      - canonical_name: "Remarks (if any)"
        synonyms: ["Remarks (if any)"]
        datatype: "string"
      - canonical_name: "PAN Number"
        synonyms: ["PAN Number"]
        datatype: "string"
      - canonical_name: "Bank Name"
        synonyms: ["Bank Name"]
        datatype: "string"
      - canonical_name: "A/c Number"
        synonyms: ["A/c Number"]
        datatype: "string"
      - canonical_name: "IFSC Code"
        synonyms: ["IFSC Code"]
        datatype: "string"
  "Master Data":
    header_row: 1
    data_start_row: 2
    columns:
      - canonical_name: "Sr No"
        synonyms: ["Sr No"]
        datatype: "integer"
        mandatory: true
      - canonical_name: "Client"
        datatype: "string"
        default_value: "Axis Bank POA"
      - canonical_name: "Month"
        synonyms: ["Month"]
        datatype: "string"
      - canonical_name: "Zone"
        synonyms: ["Zone"]
        datatype: "string"
      - canonical_name: "SOL ID"
        synonyms: ["SOL ID"]
        datatype: "string"
        mandatory: true
      - canonical_name: "BRANCH"
        synonyms: ["BRANCH"]
        datatype: "string"
        mandatory: true
      - canonical_name: "Location "
        synonyms: ["Location "]
        datatype: "string"
      - canonical_name: "State"
        synonyms: ["State"]
        datatype: "string"
      - canonical_name: "Total No.of A/cs"
        synonyms: ["Total No.of A/cs"]
        datatype: "decimal"
      - canonical_name: "Assayer Name"
        synonyms: ["Assayer Name"]
        datatype: "string"
        mandatory: true
      - canonical_name: "Assayer Code"
        synonyms: ["Assayer Code"]
        datatype: "string"
        mandatory: true
      - canonical_name: "AssayerPhone"
        synonyms: ["AssayerPhone"]
        datatype: "string"
      - canonical_name: "Assayer PAN"
        synonyms: ["Assayer PAN"]
        datatype: "string"
      - canonical_name: "Contact Person"
        synonyms: ["Contact Person"]
        datatype: "string"
      - canonical_name: "Schedule date"
        synonyms: ["Schedule date"]
        datatype: "date"
      - canonical_name: "Audit \nStatus"
        synonyms: ["Audit Status"]
        datatype: "string"
      - canonical_name: "Audit \ncompletion date"
        synonyms: ["Audit completion date"]
        datatype: "date"
      - canonical_name: "No of days \naudited "
        synonyms: ["No of days audited "]
        datatype: "decimal"
      - canonical_name: "No of days \naudited For client"
        synonyms: ["No of days audited For client"]
        datatype: "decimal"
      - canonical_name: "No of Packets \naudited"
        synonyms: ["No of Packets audited"]
        datatype: "decimal"
      - canonical_name: "Additional Packet"
        synonyms: ["Additional Packet"]
        datatype: "decimal"
      - canonical_name: "Assayer Reporting time at Branch"
        synonyms: ["Assayer Reporting time at Branch"]
        datatype: "string"
      - canonical_name: "Audit start time"
        synonyms: ["Audit start time"]
        datatype: "string"
      - canonical_name: "Audit End Time"
        synonyms: ["Audit End Time"]
        datatype: "string"
      - canonical_name: "Client fee"
        synonyms: ["Client fee"]
        datatype: "decimal"
      - canonical_name: "Additional"
        synonyms: ["Additional"]
        datatype: "decimal"
      - canonical_name: "Final Client Fees"
        synonyms: ["Final Client Fees"]
        datatype: "decimal"
      - canonical_name: "Assayer fee"
        synonyms: ["Assayer fee"]
        datatype: "decimal"
      - canonical_name: "Additional fee"
        synonyms: ["Additional fee"]
        datatype: "decimal"
      - canonical_name: "Distance"
        synonyms: ["Distance"]
        datatype: "string"
      - canonical_name: "Base Location"
        synonyms: ["Base Location"]
        datatype: "string"
      - canonical_name: "Remarks"
        synonyms: ["Remarks"]
        datatype: "string"
      - canonical_name: "Assayer fee.1"
        synonyms: ["Assayer fee.1"]
        datatype: "decimal"
      - canonical_name: "Additional fee.1"
        synonyms: ["Additional fee.1"]
        datatype: "decimal"
      - canonical_name: "Cancelled"
        synonyms: ["Cancelled"]
        datatype: "decimal"
      - canonical_name: "Error Deduciton"
        synonyms: ["Error Deduciton"]
        datatype: "decimal"
      - canonical_name: "Total"
        synonyms: ["Total"]
        datatype: "decimal"
      - canonical_name: "Audit Remarks"
        synonyms: ["Audit Remarks"]
        datatype: "string"
"""

    rbl_yaml = """
client_id: "rbl_poa"
client_display_name: "RBL Bank Muthoot Fincorp POA"
filename_pattern: "*RBL(Muthoot Fincorp)*"
active: true
sheets:
  "Payment Tracker":
    header_row: 1
    data_start_row: 2
    client_column: "client"
    s_no_column: "S.no"
    sum_columns:
      - "Total pay (Base)"
      - "Total pay"
    columns:
      - canonical_name: "S.no"
        synonyms: ["S.no"]
        datatype: "integer"
        mandatory: true
      - canonical_name: "client"
        datatype: "string"
        default_value: "RBL(muthoot)"
      - canonical_name: "Assayer Name"
        synonyms: ["Assayer Name"]
        datatype: "string"
      - canonical_name: "Assayer Code"
        synonyms: ["Assayer Code"]
        datatype: "string"
        mandatory: true
      - canonical_name: "Assayer Phone"
        synonyms: ["Assayer Phone"]
        datatype: "string"
      - canonical_name: "Location"
        synonyms: ["Location"]
        datatype: "string"
      - canonical_name: "State"
        synonyms: ["State"]
        datatype: "string"
      - canonical_name: "Zone"
        synonyms: ["Zone"]
        datatype: "string"
      - canonical_name: "Audit Month & Year"
        synonyms: ["Audit Month & Year"]
        datatype: "string"
      - canonical_name: "Type of Audit"
        synonyms: ["Type of Audit"]
        datatype: "string"
      - canonical_name: "No. of Visits"
        synonyms: ["No of visits"]
        datatype: "integer"
      - canonical_name: "Base Audit Fee"
        synonyms: ["Base Audit Fee"]
        datatype: "decimal"
      - canonical_name: "Total pay (Base)"
        synonyms: ["Total pay (Base)"]
        datatype: "decimal"
      - canonical_name: " Travel charges(If any)"
        synonyms: [" Travel charges(If any)"]
        datatype: "decimal"
      - canonical_name: "Cancelled visits"
        synonyms: ["Audit Cancellation Fees"]
        datatype: "decimal"
      - canonical_name: "Branch Cancellation Charges"
        synonyms: ["Branch Cancellation Charges"]
        datatype: "decimal"
      - canonical_name: " Andaman & Nicobar Branch Expenses"
        synonyms: [" Andaman & Nicobar Branch Expenses"]
        datatype: "decimal"
      - canonical_name: "Error Deduction"
        synonyms: ["Error Deduction"]
        datatype: "decimal"
      - canonical_name: "Total pay"
        synonyms: ["Total pay"]
        datatype: "decimal"
        mandatory: true
      - canonical_name: "Remarks (if any)"
        synonyms: ["Remarks (if any)"]
        datatype: "string"
      - canonical_name: "PAN Number"
        synonyms: ["PAN Number"]
        datatype: "string"
      - canonical_name: "Bank Name"
        synonyms: ["Bank Name"]
        datatype: "string"
      - canonical_name: "A/c Number"
        synonyms: ["A/c Number"]
        datatype: "string"
      - canonical_name: "IFSC Code"
        synonyms: ["IFSC Code"]
        datatype: "string"
  "Master Data":
    header_row: 1
    data_start_row: 2
    columns:
      - canonical_name: "Sr No"
        synonyms: ["Sr No."]
        datatype: "integer"
        mandatory: true
      - canonical_name: "Client"
        datatype: "string"
        default_value: "RBL(muthoot)"
      - canonical_name: "Month"
        synonyms: ["Month"]
        datatype: "string"
      - canonical_name: "Zone"
        synonyms: ["Zone"]
        datatype: "string"
      - canonical_name: "SOL ID"
        synonyms: ["Branch Code"]
        datatype: "string"
        mandatory: true
      - canonical_name: "BRANCH"
        synonyms: ["BranchName"]
        datatype: "string"
        mandatory: true
      - canonical_name: "Location "
        synonyms: ["Location "]
        datatype: "string"
      - canonical_name: "State"
        synonyms: ["State"]
        datatype: "string"
      - canonical_name: "Total No.of A/cs"
        synonyms: ["Total pouches suggested for audit"]
        datatype: "decimal"
      - canonical_name: "Assayer Name"
        synonyms: ["Assayer Name"]
        datatype: "string"
        mandatory: true
      - canonical_name: "Assayer Code"
        synonyms: ["Assayer Code"]
        datatype: "string"
        mandatory: true
      - canonical_name: "AssayerPhone"
        synonyms: ["AssayerPhone"]
        datatype: "string"
      - canonical_name: "Assayer PAN"
        synonyms: ["Assayer PAN"]
        datatype: "string"
      - canonical_name: "Contact Person"
        synonyms: ["Process Manager"]
        datatype: "string"
      - canonical_name: "Schedule date"
        synonyms: ["Audit schedule date"]
        datatype: "date"
      - canonical_name: "Audit \nStatus"
        synonyms: ["Audit Status"]
        datatype: "string"
      - canonical_name: "Audit \ncompletion date"
        synonyms: ["End Date"]
        datatype: "date"
      - canonical_name: "No of days \naudited "
        synonyms: ["No. of Visit"]
        datatype: "decimal"
      - canonical_name: "No of days \naudited For client"
        synonyms: ["No. of Visit"]
        datatype: "decimal"
      - canonical_name: "No of Packets \naudited"
        synonyms: ["No of Packets audited"]
        datatype: "decimal"
      - canonical_name: "Additional Packet"
        synonyms: ["Additional Packet"]
        datatype: "decimal"
      - canonical_name: "Assayer Reporting time at Branch"
        synonyms: ["Assayer Reporting time at Branch"]
        datatype: "string"
      - canonical_name: "Audit start time"
        synonyms: ["Audit start time"]
        datatype: "string"
      - canonical_name: "Audit End Time"
        synonyms: ["Audit End Time"]
        datatype: "string"
      - canonical_name: "Client fee"
        synonyms: ["Client Fees"]
        datatype: "decimal"
      - canonical_name: "Additional"
        synonyms: ["Additional"]
        datatype: "decimal"
      - canonical_name: "Final Client Fees"
        synonyms: ["Final Client Fees"]
        datatype: "decimal"
      - canonical_name: "Assayer fee"
        synonyms: ["Assayer Fees"]
        datatype: "decimal"
      - canonical_name: "Additional fee"
        synonyms: ["Additional fee"]
        datatype: "decimal"
      - canonical_name: "Distance"
        synonyms: ["Distance"]
        datatype: "string"
      - canonical_name: "Base Location"
        synonyms: ["Assayer Base location"]
        datatype: "string"
      - canonical_name: "Remarks"
        synonyms: ["Remarks"]
        datatype: "string"
      - canonical_name: "Assayer fee.1"
        synonyms: ["Assayer Fees.1"]
        datatype: "decimal"
      - canonical_name: "Additional fee.1"
        synonyms: ["Additional fee.1"]
        datatype: "decimal"
      - canonical_name: "Cancelled"
        synonyms: ["Cancellation"]
        datatype: "decimal"
      - canonical_name: "Error Deduciton"
        synonyms: ["Error Deduction"]
        datatype: "decimal"
      - canonical_name: "Total"
        synonyms: ["Total"]
        datatype: "decimal"
      - canonical_name: "Audit Remarks"
        synonyms: ["Other Remarks"]
        datatype: "string"
"""

    axis_config_path = os.path.join(test_workspace, "config", "schemas", "axis_poa.yaml")
    with open(axis_config_path, "w") as f:
        f.write(axis_yaml)

    rbl_config_path = os.path.join(test_workspace, "config", "schemas", "rbl_poa.yaml")
    with open(rbl_config_path, "w") as f:
        f.write(rbl_yaml)

    output_xlsx = os.path.join(test_workspace, "Feb'26 consolidated.xlsx")

    # 2. Run E2E pipeline
    execute_e2e_consolidation(test_workspace, output_xlsx)

    # 3. Assertions checking output workbook is generated correctly
    assert os.path.exists(output_xlsx)
    
    # Check sheet contents
    df_pt = pd.read_excel(output_xlsx, sheet_name="Payment Tracker")
    df_md = pd.read_excel(output_xlsx, sheet_name="Master Data")

    # Clean rows count should match standalone (1 Axis PT + 1 RBL PT = 2 clean PT rows)
    # Pandas drops the spacer and formula-only totals rows as they have only NaN/formula values
    assert len(df_pt) == 2

    # Assert using openpyxl that the dynamic totals row is written at row 5
    import openpyxl
    wb = openpyxl.load_workbook(output_xlsx)
    ws = wb["Payment Tracker"]
    # Row 4 is empty spacer row
    assert ws.cell(row=4, column=1).value is None
    # Row 5 has dynamic totals SUM formulas
    cols = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    assert "Total pay" in cols
    tot_pay_idx = cols.index("Total pay") + 1
    col_letter = openpyxl.utils.get_column_letter(tot_pay_idx)
    assert ws.cell(row=5, column=tot_pay_idx).value == f"=SUM({col_letter}2:{col_letter}4)"
    
    # First row is Axis Bank
    assert df_pt.iloc[0]["client"] == "Axis Bank POA"
    assert df_pt.iloc[0]["Total pay"] == 13500.0
    
    # Second row is RBL
    assert df_pt.iloc[1]["client"] == "RBL(muthoot)"
    assert df_pt.iloc[1]["Total pay"] == 11500.0

    # Master Data count: 1 Axis MD + 1 RBL MD = 2 clean MD rows
    assert len(df_md) == 2
    assert df_md.iloc[0]["Client"] == "Axis Bank POA"
    assert df_md.iloc[0]["Location "] == "Delhi" # Location seeded from State
    assert df_md.iloc[1]["Client"] == "RBL(muthoot)"
